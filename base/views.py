# base/views.py - Vues complètes du module de base AVEC EXPORT

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, F
from django.http import HttpResponse
from datetime import datetime, date, timedelta
from django.utils import timezone
from io import BytesIO
import csv

from .models import Client, Fournisseur
from .forms import FormulaireClient, FormulaireFournisseur
from stock.models import Produit, MouvementStock
from ventes.models import CommandeVente, Facture

# ========== TABLEAU DE BORD ==========

@login_required
def tableau_bord(request):
    """Vue pour le tableau de bord principal"""
    from stock.models import Stock
    from django.db.models import Sum, Count, Q, F
    from datetime import datetime, timedelta
    
    aujourdhui = datetime.now().date()
    debut_mois = aujourdhui.replace(day=1)
    
    # ==================== STATISTIQUES GÉNÉRALES ====================
    
    # Clients actifs
    total_clients = Client.objects.filter(est_actif=True).count()
    
    # Produits actifs
    total_produits = Produit.objects.filter(est_actif=True).count()
    
    # Commandes en cours (confirmées mais non livrées)
    commandes_en_cours = CommandeVente.objects.filter(
        statut__in=['CONFIRME', 'EN_PREPARATION']
    ).count()
    
    # ==================== GESTION DU STOCK ====================
    
    # Produits en stock bas (en dessous du seuil de réapprovisionnement)
    produits_stock_bas = Stock.objects.filter(
        produit__est_actif=True
    ).annotate(
        est_stock_bas=Q(quantite__lte=F('produit__seuil_reapprovisionnement'))
    ).filter(
        est_stock_bas=True
    ).values('produit').distinct().count()
    
    # Produits en rupture de stock (quantité = 0)
    produits_rupture = Stock.objects.filter(
        produit__est_actif=True,
        quantite=0
    ).count()
    
    # Produits sans entrée de stock
    produits_avec_stock = Stock.objects.filter(
        produit__est_actif=True
    ).values_list('produit_id', flat=True)
    
    produits_sans_stock = Produit.objects.filter(
        est_actif=True
    ).exclude(
        id__in=produits_avec_stock
    ).count()
    
    # Total des produits nécessitant attention (stock bas + rupture + sans stock)
    total_produits_attention = produits_stock_bas + produits_rupture + produits_sans_stock
    
    # ==================== STATISTIQUES VENTES ====================
    
    # Ventes du mois en cours
    ventes_mois = Facture.objects.filter(
        date_facture__gte=debut_mois,
        statut__in=['PAYEE', 'EN_ATTENTE', 'PARTIELLE']
    ).aggregate(
        total=Sum('total')
    )['total'] or 0
    
    # Nombre de commandes du mois
    commandes_mois = CommandeVente.objects.filter(
        date_commande__gte=debut_mois
    ).count()
    
    # Factures impayées
    factures_impayees = Facture.objects.filter(
        statut__in=['EN_ATTENTE', 'EN_RETARD', 'PARTIELLE']
    ).count()
    
    # Montant total des factures impayées
    montant_impaye = Facture.objects.filter(
        statut__in=['EN_ATTENTE', 'EN_RETARD', 'PARTIELLE']
    ).aggregate(
        total=Sum('total')
    )['total'] or 0
    
    # ==================== DONNÉES POUR LES TABLEAUX ====================
    
    # Commandes récentes (10 dernières)
    commandes_recentes = CommandeVente.objects.select_related(
        'client'
    ).order_by('-date_creation')[:10]
    
    # Factures récentes (10 dernières)
    factures_recentes = Facture.objects.select_related(
        'client'
    ).order_by('-date_creation')[:10]
    
    # Produits nécessitant réapprovisionnement (5 premiers)
    produits_reappro = Stock.objects.filter(
        produit__est_actif=True,
        quantite__lte=F('produit__seuil_reapprovisionnement')
    ).select_related('produit').order_by('quantite')[:5]
    
    # ==================== ACTIVITÉS RÉCENTES ====================
    
    # Dernières commandes créées (pour la timeline)
    dernieres_commandes = CommandeVente.objects.select_related(
        'client'
    ).order_by('-date_creation')[:3]
    
    # Derniers clients ajoutés
    derniers_clients = Client.objects.filter(
        est_actif=True
    ).order_by('-id')[:3]
    
    # ==================== STATISTIQUES COMPARATIVES ====================
    
    # Ventes du mois précédent pour comparaison
    debut_mois_precedent = (debut_mois - timedelta(days=1)).replace(day=1)
    fin_mois_precedent = debut_mois - timedelta(days=1)
    
    ventes_mois_precedent = Facture.objects.filter(
        date_facture__gte=debut_mois_precedent,
        date_facture__lte=fin_mois_precedent,
        statut__in=['PAYEE', 'EN_ATTENTE', 'PARTIELLE']
    ).aggregate(
        total=Sum('total')
    )['total'] or 0
    
    # Calcul du taux de croissance
    if ventes_mois_precedent > 0:
        taux_croissance = ((ventes_mois - ventes_mois_precedent) / ventes_mois_precedent) * 100
    else:
        taux_croissance = 100 if ventes_mois > 0 else 0
    
    # ==================== CONTEXTE ====================
    
    contexte = {
        # Statistiques générales
        'total_clients': total_clients,
        'total_produits': total_produits,
        'commandes_en_cours': commandes_en_cours,
        'produits_stock_bas': total_produits_attention,
        
        # Détails stock
        'produits_rupture': produits_rupture,
        'produits_sans_stock': produits_sans_stock,
        'produits_reappro': produits_reappro,
        
        # Ventes
        'ventes_mois': ventes_mois,
        'commandes_mois': commandes_mois,
        'factures_impayees': factures_impayees,
        'montant_impaye': montant_impaye,
        'taux_croissance': taux_croissance,
        
        # Tableaux
        'commandes_recentes': commandes_recentes,
        'factures_recentes': factures_recentes,
        
        # Activités
        'dernieres_commandes': dernieres_commandes,
        'derniers_clients': derniers_clients,
        
        # Date
        'aujourdhui': aujourdhui,
    }
    
    return render(request, 'tableau_bord.jinja', contexte)

# ========== GESTION DES CLIENTS ==========

@login_required
def liste_clients(request):
    """Vue pour afficher la liste des clients"""
    recherche = request.GET.get('recherche', '')
    clients = Client.objects.filter(est_actif=True)
    
    if recherche:
        clients = clients.filter(
            Q(code__icontains=recherche) |
            Q(nom__icontains=recherche) |
            Q(email__icontains=recherche)
        )
    
    clients = clients.order_by('-date_creation')
    
    # Calculer le solde pour chaque client
    for client in clients:
        solde = Facture.objects.filter(client=client).aggregate(
            total_solde=Sum(F('total') - F('montant_paye'))
        )['total_solde'] or 0
        client.solde_actuel = solde
    
    contexte = {
        'clients': clients,
        'recherche': recherche
    }
    
    return render(request, 'base/liste_clients.jinja', contexte)

@login_required
def creer_client(request):
    """Vue pour créer un nouveau client"""
    if request.method == 'POST':
        formulaire = FormulaireClient(request.POST)
        if formulaire.is_valid():
            formulaire.save()
            messages.success(request, 'Client créé avec succès!')
            return redirect('liste_clients')
    else:
        formulaire = FormulaireClient()
    
    contexte = {
        'formulaire': formulaire,
        'action': 'Créer'
    }
    
    return render(request, 'base/formulaire_client.jinja', contexte)

@login_required
def modifier_client(request, pk):
    """Vue pour modifier un client existant"""
    client = get_object_or_404(Client, pk=pk)
    
    if request.method == 'POST':
        formulaire = FormulaireClient(request.POST, instance=client)
        if formulaire.is_valid():
            formulaire.save()
            messages.success(request, 'Client modifié avec succès!')
            return redirect('liste_clients')
    else:
        formulaire = FormulaireClient(instance=client)
    
    contexte = {
        'formulaire': formulaire,
        'action': 'Modifier',
        'client': client
    }
    
    return render(request, 'base/formulaire_client.jinja', contexte)

@login_required
def details_client(request, pk):
    """Vue pour afficher les détails d'un client"""
    client = get_object_or_404(Client, pk=pk)
    
    # Récupérer les commandes et factures
    commandes = CommandeVente.objects.filter(client=client).order_by('-date_creation')[:10]
    factures = Facture.objects.filter(client=client).order_by('-date_creation')[:10]

    # Calculer le solde total du client (total - montant_paye)
    solde_client = Facture.objects.filter(client=client).aggregate(
        total_solde=Sum(F('total') - F('montant_paye'))
    )['total_solde'] or 0
    
    # Calculer les totaux
    total_ventes = CommandeVente.objects.filter(client=client).aggregate(
        total=Sum('total')
    )['total'] or 0
    
    factures_impayees = Facture.objects.filter(
        client=client,
        statut__in=['BROUILLON', 'ENVOYEE']
    ).aggregate(total=Sum('total'))['total'] or 0
    
    contexte = {
        'client': client,
        'commandes': commandes,
        'factures': factures,
        'solde_client': solde_client,
        'total_ventes': total_ventes,
        'factures_impayees': factures_impayees,
    }
    
    return render(request, 'base/details_client.jinja', contexte)

@login_required
def supprimer_client(request, pk):
    """Vue pour désactiver un client"""
    client = get_object_or_404(Client, pk=pk)
    
    if request.method == 'POST':
        client.est_actif = False
        client.save()
        messages.success(request, 'Client désactivé avec succès!')
        return redirect('liste_clients')
    
    contexte = {'client': client}
    return render(request, 'base/confirmer_suppression_client.jinja', contexte)

# ========== GESTION DES FOURNISSEURS ==========

@login_required
def liste_fournisseurs(request):
    """Vue pour afficher la liste des fournisseurs"""
    recherche = request.GET.get('recherche', '')
    fournisseurs = Fournisseur.objects.filter(est_actif=True)
    
    if recherche:
        fournisseurs = fournisseurs.filter(
            Q(code__icontains=recherche) |
            Q(nom__icontains=recherche) |
            Q(email__icontains=recherche)
        )
    
    fournisseurs = fournisseurs.order_by('-date_creation')
    
    # Statistiques pour la liste
    fournisseurs_actifs = fournisseurs.filter(est_actif=True).count()
    
    contexte = {
        'fournisseurs': fournisseurs,
        'recherche': recherche,
        'fournisseurs_actifs': fournisseurs_actifs,
    }
    
    return render(request, 'base/liste_fournisseurs.jinja', contexte)

@login_required
def creer_fournisseur(request):
    """Vue pour créer un nouveau fournisseur"""
    if request.method == 'POST':
        formulaire = FormulaireFournisseur(request.POST)
        if formulaire.is_valid():
            formulaire.save()
            messages.success(request, 'Fournisseur créé avec succès!')
            return redirect('liste_fournisseurs')
    else:
        formulaire = FormulaireFournisseur()
    
    # Statistiques pour la sidebar
    fournisseurs_actifs = Fournisseur.objects.filter(est_actif=True).count()
    delai_moyen = Fournisseur.objects.filter(est_actif=True).aggregate(
        moyenne=Sum('delai_paiement')
    )['moyenne'] or 30
    
    contexte = {
        'formulaire': formulaire,
        'action': 'Créer',
        'fournisseurs_actifs': fournisseurs_actifs,
        'delai_moyen': delai_moyen / fournisseurs_actifs if fournisseurs_actifs > 0 else 30,
    }
    
    return render(request, 'base/formulaire_fournisseur.jinja', contexte)

@login_required
def modifier_fournisseur(request, pk):
    """Vue pour modifier un fournisseur existant"""
    fournisseur = get_object_or_404(Fournisseur, pk=pk)
    
    if request.method == 'POST':
        formulaire = FormulaireFournisseur(request.POST, instance=fournisseur)
        if formulaire.is_valid():
            formulaire.save()
            messages.success(request, 'Fournisseur modifié avec succès!')
            return redirect('liste_fournisseurs')
    else:
        formulaire = FormulaireFournisseur(instance=fournisseur)
    
    # Statistiques pour la sidebar
    fournisseurs_actifs = Fournisseur.objects.filter(est_actif=True).count()
    delai_moyen = Fournisseur.objects.filter(est_actif=True).aggregate(
        moyenne=Sum('delai_paiement')
    )['moyenne'] or 30
    
    contexte = {
        'formulaire': formulaire,
        'action': 'Modifier',
        'fournisseur': fournisseur,
        'fournisseurs_actifs': fournisseurs_actifs,
        'delai_moyen': delai_moyen / fournisseurs_actifs if fournisseurs_actifs > 0 else 30,
    }
    
    return render(request, 'base/formulaire_fournisseur.jinja', contexte)

@login_required
def details_fournisseur(request, pk):
    """Vue pour afficher les détails d'un fournisseur"""
    from achats.models import CommandeAchat
    
    fournisseur = get_object_or_404(Fournisseur, pk=pk)
    
    # Récupérer les commandes
    commandes = CommandeAchat.objects.filter(
        fournisseur=fournisseur
    ).order_by('-date_creation')[:10]
    
    # ==================== CALCUL DES DÉLAIS ====================
    aujourdhui = date.today()
    
    for commande in commandes:
        if commande.date_livraison_prevue:
            # Convertir en date si c'est un datetime
            if hasattr(commande.date_livraison_prevue, 'date'):
                date_livraison = commande.date_livraison_prevue.date()
            else:
                date_livraison = commande.date_livraison_prevue
            
            # Calculer les jours restants
            jours_restants = (date_livraison - aujourdhui).days
            commande.jours_restants = jours_restants
            
            # Déterminer le statut du délai
            if jours_restants < 0:
                commande.delai_statut = 'late'  # En retard
            elif jours_restants <= 3:
                commande.delai_statut = 'soon'  # Bientôt
            else:
                commande.delai_statut = 'ok'    # OK
        else:
            commande.jours_restants = None
            commande.delai_statut = None
    
    # Calculer les statistiques
    total_achats = CommandeAchat.objects.filter(
        fournisseur=fournisseur
    ).aggregate(total=Sum('total'))['total'] or 0
    
    # Commandes du mois en cours
    debut_mois = aujourdhui.replace(day=1)
    commandes_ce_mois = CommandeAchat.objects.filter(
        fournisseur=fournisseur,
        date_creation__gte=debut_mois
    ).count()
    
    # Récupérer les factures (si le modèle existe)
    try:
        from comptabilite.models import FactureFournisseur
        factures = FactureFournisseur.objects.filter(
            fournisseur=fournisseur
        ).order_by('-date_facture')[:10]
    except ImportError:
        factures = []
    
    contexte = {
        'fournisseur': fournisseur,
        'commandes': commandes,
        'factures': factures,
        'total_achats': total_achats,
        'commandes_ce_mois': commandes_ce_mois,
    }
    
    return render(request, 'base/details_fournisseur.jinja', contexte)

@login_required
def supprimer_fournisseur(request, pk):
    """Vue pour désactiver un fournisseur"""
    fournisseur = get_object_or_404(Fournisseur, pk=pk)
    
    if request.method == 'POST':
        fournisseur.est_actif = False
        fournisseur.save()
        messages.success(request, 'Fournisseur désactivé avec succès!')
        return redirect('liste_fournisseurs')
    
    contexte = {'fournisseur': fournisseur}
    return render(request, 'base/confirmer_suppression_fournisseur.jinja', contexte)


# ========== EXPORT DES FOURNISSEURS ==========

@login_required
def exporter_fournisseurs(request):
    """
    Exporte la liste des fournisseurs en Excel, CSV ou PDF
    """
    format_export = request.GET.get('format', 'excel')
    recherche = request.GET.get('recherche', '')
    ids = request.GET.get('ids', '')  # Pour export sélection
    
    # Récupérer les fournisseurs
    fournisseurs = Fournisseur.objects.filter(est_actif=True)
    
    # Filtrer par recherche si spécifiée
    if recherche:
        fournisseurs = fournisseurs.filter(
            Q(code__icontains=recherche) |
            Q(nom__icontains=recherche) |
            Q(email__icontains=recherche)
        )
    
    # Filtrer par IDs si sélection spécifique
    if ids:
        id_list = [int(id) for id in ids.split(',')]
        fournisseurs = fournisseurs.filter(pk__in=id_list)
    
    fournisseurs = fournisseurs.order_by('nom')
    
    # ==================== EXPORT EXCEL ====================
    if format_export == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Fournisseurs"
            
            # En-têtes
            headers = ['Code', 'Nom', 'Email', 'Téléphone', 'Ville', 'Pays', 'Délai paiement', 'Type', 'Statut']
            ws.append(headers)
            
            # Style des en-têtes
            header_fill = PatternFill(start_color="0ea5e9", end_color="0ea5e9", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Données
            for fournisseur in fournisseurs:
                ws.append([
                    fournisseur.code,
                    fournisseur.nom,
                    fournisseur.email or '-',
                    fournisseur.telephone or '-',
                    fournisseur.ville or '-',
                    fournisseur.pays or '-',
                    f"{fournisseur.delai_paiement or 0} jours",
                    fournisseur.get_type_fournisseur_display() if hasattr(fournisseur, 'type_fournisseur') else '-',
                    'Actif' if fournisseur.est_actif else 'Inactif'
                ])
            
            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Créer la réponse HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="fournisseurs.xlsx"'
            wb.save(response)
            return response
            
        except ImportError:
            messages.error(request, 'La bibliothèque openpyxl n\'est pas installée. Utilisez: pip install openpyxl')
            return redirect('liste_fournisseurs')
    
    # ==================== EXPORT CSV ====================
    elif format_export == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="fournisseurs.csv"'
        
        # BOM UTF-8 pour Excel
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow(['Code', 'Nom', 'Email', 'Téléphone', 'Ville', 'Pays', 'Délai paiement', 'Type', 'Statut'])
        
        for fournisseur in fournisseurs:
            writer.writerow([
                fournisseur.code,
                fournisseur.nom,
                fournisseur.email or '-',
                fournisseur.telephone or '-',
                fournisseur.ville or '-',
                fournisseur.pays or '-',
                f"{fournisseur.delai_paiement or 0} jours",
                fournisseur.get_type_fournisseur_display() if hasattr(fournisseur, 'type_fournisseur') else '-',
                'Actif' if fournisseur.est_actif else 'Inactif'
            ])
        
        return response
    
    # ==================== EXPORT PDF ====================
    elif format_export == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
            elements = []
            styles = getSampleStyleSheet()
            
            # Titre
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#0ea5e9'),
                spaceAfter=30,
            )
            title = Paragraph("Liste des Fournisseurs", title_style)
            elements.append(title)
            
            # Date
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
            )
            date_text = Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style)
            elements.append(date_text)
            elements.append(Spacer(1, 20))
            
            # Données du tableau
            data = [['Code', 'Nom', 'Email', 'Téléphone', 'Ville', 'Pays', 'Délai', 'Statut']]
            
            for fournisseur in fournisseurs:
                data.append([
                    fournisseur.code[:10],
                    fournisseur.nom[:20],
                    (fournisseur.email or '-')[:25],
                    (fournisseur.telephone or '-')[:15],
                    (fournisseur.ville or '-')[:15],
                    (fournisseur.pays or '-')[:15],
                    f"{fournisseur.delai_paiement or 0}j",
                    'Actif' if fournisseur.est_actif else 'Inactif'
                ])
            
            # Créer le tableau
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0ea5e9')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Contenu
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # Lignes alternées
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ]))
            
            elements.append(table)
            
            # Footer
            elements.append(Spacer(1, 20))
            footer_text = Paragraph(
                f"Total: {len(fournisseurs)} fournisseur(s)",
                ParagraphStyle('Footer', parent=styles['Normal'], fontSize=10, textColor=colors.grey)
            )
            elements.append(footer_text)
            
            # Construire le PDF
            doc.build(elements)
            
            pdf = buffer.getvalue()
            buffer.close()
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="fournisseurs.pdf"'
            response.write(pdf)
            return response
            
        except ImportError:
            messages.error(request, 'La bibliothèque reportlab n\'est pas installée. Utilisez: pip install reportlab')
            return redirect('liste_fournisseurs')
    
    else:
        messages.error(request, 'Format d\'export non reconnu')
        return redirect('liste_fournisseurs')


@login_required
def action_groupee_fournisseurs(request):
    """
    Gère les actions groupées sur les fournisseurs (activer, désactiver, supprimer)
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        fournisseur_ids = request.POST.getlist('fournisseur_ids')
        
        if not fournisseur_ids:
            messages.error(request, 'Aucun fournisseur sélectionné')
            return redirect('liste_fournisseurs')
        
        fournisseurs = Fournisseur.objects.filter(pk__in=fournisseur_ids)
        count = fournisseurs.count()
        
        if action == 'activate':
            fournisseurs.update(est_actif=True)
            messages.success(request, f'{count} fournisseur(s) activé(s) avec succès')
        
        elif action == 'deactivate':
            fournisseurs.update(est_actif=False)
            messages.success(request, f'{count} fournisseur(s) désactivé(s) avec succès')
        
        elif action == 'delete':
            fournisseurs.update(est_actif=False)  # Soft delete
            messages.success(request, f'{count} fournisseur(s) supprimé(s) avec succès')
        
        else:
            messages.error(request, 'Action non reconnue')
        
        return redirect('liste_fournisseurs')
    
    return redirect('liste_fournisseurs')
# Fonctions à ajouter dans base/views.py pour l'export des clients

@login_required
def exporter_clients(request):
    """
    Exporte la liste des clients en Excel, CSV ou PDF
    """
    format_export = request.GET.get('format', 'excel')
    recherche = request.GET.get('recherche', '')
    ids = request.GET.get('ids', '')  # Pour export sélection
    
    # Récupérer les clients
    clients = Client.objects.filter(est_actif=True)
    
    # Filtrer par recherche si spécifiée
    if recherche:
        clients = clients.filter(
            Q(code__icontains=recherche) |
            Q(nom__icontains=recherche) |
            Q(email__icontains=recherche)
        )
    
    # Filtrer par IDs si sélection spécifique
    if ids:
        id_list = [int(id) for id in ids.split(',')]
        clients = clients.filter(pk__in=id_list)
    
    clients = clients.order_by('nom')
    
    # Calculer le solde pour chaque client
    for client in clients:
        # Calculer le solde du client (somme des factures impayées)
        solde = Facture.objects.filter(client=client).aggregate(
            total_solde=Sum(F('total') - F('montant_paye'))
        )['total_solde'] or 0
        client.solde_calcule = solde
    
    # ==================== EXPORT EXCEL ====================
    if format_export == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Clients"
            
            # En-têtes
            headers = ['Code', 'Nom', 'Email', 'Téléphone', 'Ville', 'Pays', 'Limite crédit', 'Solde actuel', 'Type', 'Statut']
            ws.append(headers)
            
            # Style des en-têtes
            header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Données
            for client in clients:
                ws.append([
                    client.code,
                    client.nom,
                    client.email or '-',
                    client.telephone or '-',
                    client.ville or '-',
                    client.pays or '-',
                    client.limite_credit,
                    client.solde_calcule,
                    client.get_type_client_display() if hasattr(client, 'type_client') else '-',
                    'Actif' if client.est_actif else 'Inactif'
                ])
            
            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Créer la réponse HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="clients.xlsx"'
            wb.save(response)
            return response
            
        except ImportError:
            messages.error(request, 'La bibliothèque openpyxl n\'est pas installée. Utilisez: pip install openpyxl')
            return redirect('liste_clients')
    
    # ==================== EXPORT CSV ====================
    elif format_export == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="clients.csv"'
        
        # BOM UTF-8 pour Excel
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow(['Code', 'Nom', 'Email', 'Téléphone', 'Ville', 'Pays', 'Limite crédit', 'Solde actuel', 'Type', 'Statut'])
        
        for client in clients:
            writer.writerow([
                client.code,
                client.nom,
                client.email or '-',
                client.telephone or '-',
                client.ville or '-',
                client.pays or '-',
                client.limite_credit,
                client.solde_calcule,
                client.get_type_client_display() if hasattr(client, 'type_client') else '-',
                'Actif' if client.est_actif else 'Inactif'
            ])
        
        return response
    
    # ==================== EXPORT PDF ====================
    elif format_export == 'pdf':
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
            elements = []
            styles = getSampleStyleSheet()
            
            # Titre
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1e40af'),
                spaceAfter=30,
            )
            title = Paragraph("Liste des Clients", title_style)
            elements.append(title)
            
            # Date
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
            )
            date_text = Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style)
            elements.append(date_text)
            elements.append(Spacer(1, 20))
            
            # Données du tableau
            data = [['Code', 'Nom', 'Email', 'Téléphone', 'Ville', 'Limite', 'Solde', 'Statut']]
            
            for client in clients:
                data.append([
                    client.code[:10],
                    client.nom[:20],
                    (client.email or '-')[:25],
                    (client.telephone or '-')[:15],
                    (client.ville or '-')[:15],
                    f"{client.limite_credit:,.0f}",
                    f"{client.solde_calcule:,.0f}",
                    'Actif' if client.est_actif else 'Inactif'
                ])
            
            # Créer le tableau
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Contenu
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # Lignes alternées
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ]))
            
            elements.append(table)
            
            # Footer
            elements.append(Spacer(1, 20))
            footer_text = Paragraph(
                f"Total: {len(clients)} client(s)",
                ParagraphStyle('Footer', parent=styles['Normal'], fontSize=10, textColor=colors.grey)
            )
            elements.append(footer_text)
            
            # Construire le PDF
            doc.build(elements)
            
            pdf = buffer.getvalue()
            buffer.close()
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="clients.pdf"'
            response.write(pdf)
            return response
            
        except ImportError:
            messages.error(request, 'La bibliothèque reportlab n\'est pas installée. Utilisez: pip install reportlab')
            return redirect('liste_clients')
    
    else:
        messages.error(request, 'Format d\'export non reconnu')
        return redirect('liste_clients')


@login_required
def action_groupee_clients(request):
    """
    Gère les actions groupées sur les clients (activer, désactiver, supprimer)
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        client_ids = request.POST.getlist('client_ids')
        
        if not client_ids:
            messages.error(request, 'Aucun client sélectionné')
            return redirect('liste_clients')
        
        clients = Client.objects.filter(pk__in=client_ids)
        count = clients.count()
        
        if action == 'activate':
            clients.update(est_actif=True)
            messages.success(request, f'{count} client(s) activé(s) avec succès')
        
        elif action == 'deactivate':
            clients.update(est_actif=False)
            messages.success(request, f'{count} client(s) désactivé(s) avec succès')
        
        elif action == 'delete':
            clients.update(est_actif=False)  # Soft delete
            messages.success(request, f'{count} client(s) supprimé(s) avec succès')
        
        else:
            messages.error(request, 'Action non reconnue')
        
        return redirect('liste_clients')
    
    return redirect('liste_clients')