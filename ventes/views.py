# ventes/views.py - Vues complètes du module ventes

from html import unescape
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum
from datetime import datetime, timedelta
from datetime import date
from django.utils import timezone 
from decimal import Decimal, InvalidOperation  # AJOUT: Import Decimal
from .models import CommandeVente, LigneCommandeVente, Facture
from stock.models import Produit, MouvementStock, Entrepot  # Entrepot importé d'ici
from base.models import Client
from django.core.mail import EmailMessage
from django.conf import settings
from io import BytesIO

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm

import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ========== GESTION DES COMMANDES DE VENTE ==========

@login_required
def liste_commandes_vente(request):
    """Vue pour afficher la liste des commandes de vente"""
    filtre_statut = request.GET.get('statut', '')
    recherche = request.GET.get('recherche', '')
    
    commandes = CommandeVente.objects.select_related('client', 'entrepot', 'cree_par')
    
    if filtre_statut:
        commandes = commandes.filter(statut=filtre_statut)
    
    if recherche:
        commandes = commandes.filter(
            Q(numero_commande__icontains=recherche) |
            Q(client__nom__icontains=recherche)
        )
    
    commandes = commandes.order_by('-date_creation')

    # Calculer les statistiques
    aujourdhui = timezone.now().date()
    total_montant = sum([c.total for c in commandes if c.total])
    commandes_en_retard = commandes.filter(
        date_livraison__lt=aujourdhui, 
        statut__in=['BROUILLON', 'CONFIRME']
    ).count()
    commandes_ce_mois = commandes.filter(
        date_creation__year=aujourdhui.year,
        date_creation__month=aujourdhui.month
    ).count()
    
    contexte = {
        'commandes': commandes,
        'filtre_statut': filtre_statut,
        'recherche': recherche,
        'statuts': CommandeVente.STATUTS,
        'aujourdhui': aujourdhui,  # Ajouté
        'now': timezone.now(),  # Ajouté - pour utiliser dans le template
        'total_montant': total_montant or 0,
        'commandes_en_retard': commandes_en_retard,
        'commandes_ce_mois': commandes_ce_mois,
        'variation_mois': 0,  # À calculer si nécessaire
        'commandes_attente': commandes.filter(statut='BROUILLON').count(),
        'commandes_confirmees': commandes.filter(statut='CONFIRME').count(),
        'commandes_expediees': commandes.filter(statut='EXPEDIE').count(),
    }
    
    return render(request, 'ventes/liste_commandes.jinja', contexte)


@login_required
@transaction.atomic
def creer_commande_vente(request):
    """Vue pour créer une nouvelle commande de vente"""
    # Supprimez les imports conditionnels ici
    
    if request.method == 'POST':
        # Récupérer les données du formulaire
        client_id = request.POST.get('client')
        date_livraison = request.POST.get('date_livraison')
        entrepot_id = request.POST.get('entrepot')
        notes = request.POST.get('notes', '')
        
        # Générer le numéro de commande
        aujourdhui = datetime.now()
        prefixe = f"CV{aujourdhui.strftime('%Y%m')}"
        derniere_commande = CommandeVente.objects.filter(
            numero_commande__startswith=prefixe
        ).order_by('-numero_commande').first()
        
        if derniere_commande:
            dernier_numero = int(derniere_commande.numero_commande[-4:])
            nouveau_numero = dernier_numero + 1
        else:
            nouveau_numero = 1
        
        numero_commande = f"{prefixe}{nouveau_numero:04d}"
        
        # Créer la commande
        commande = CommandeVente.objects.create(
            numero_commande=numero_commande,
            client_id=client_id,
            date_livraison=date_livraison,
            entrepot_id=entrepot_id,
            notes=notes,
            cree_par=request.user,
            statut='BROUILLON'
        )
        
        # Créer les lignes de commande
        produits_ids = request.POST.getlist('produit[]')
        quantites = request.POST.getlist('quantite[]')
        prix = request.POST.getlist('prix_unitaire[]')
        
        for i, produit_id in enumerate(produits_ids):
            if produit_id and quantites[i]:
                produit = Produit.objects.get(pk=produit_id)
                LigneCommandeVente.objects.create(
                    commande=commande,
                    produit_id=produit_id,
                    quantite=int(quantites[i]),
                    prix_unitaire=float(prix[i]) if prix[i] else produit.prix_vente,
                    taux_tva=produit.taux_tva,
                    remise=0
                )
        
        # Calculer les totaux
        commande.calculer_totaux()
        
        messages.success(request, f'Commande {numero_commande} créée avec succès!')
        return redirect('ventes:details_commande_vente', pk=commande.pk)
    
    # GET - Afficher le formulaire
    # Utilisez les imports globaux
    clients = Client.objects.filter(est_actif=True).order_by('nom')
    entrepots = Entrepot.objects.filter(est_actif=True).order_by('nom')
    produits = Produit.objects.filter(est_actif=True).order_by('nom')
    
    contexte = {
        'clients': clients,
        'entrepots': entrepots,
        'produits': produits,
        'action': 'Créer',
        'date_livraison_defaut': (datetime.now().date() + timedelta(days=7)).strftime('%Y-%m-%d')
    }
    
    return render(request, 'ventes/formulaire_commande.jinja', contexte)


@login_required
def details_commande_vente(request, pk):
    """Vue pour afficher les détails d'une commande de vente"""
    commande = get_object_or_404(
        CommandeVente.objects.select_related('client', 'entrepot', 'cree_par'),
        pk=pk
    )
    lignes = commande.lignecommandevente_set.select_related('produit').all()

    # Ajoutez cette ligne pour obtenir la date du jour
    aujourdhui = timezone.now().date()
    
    contexte = {
        'commande': commande,
        'lignes': lignes,
        'aujourdhui': aujourdhui,
    }
    
    return render(request, 'ventes/details_commande.jinja', contexte)


@login_required
@transaction.atomic
def modifier_commande_vente(request, pk):
    """Vue pour modifier une commande de vente"""
    commande = get_object_or_404(CommandeVente, pk=pk)
    
    if commande.statut not in ['BROUILLON', 'CONFIRME']:
        messages.error(request, 'Cette commande ne peut plus être modifiée.')
        return redirect('ventes:details_commande_vente', pk=pk)
    
    if request.method == 'POST':
        # Mettre à jour la commande
        commande.client_id = request.POST.get('client')
        commande.date_livraison = request.POST.get('date_livraison')
        commande.entrepot_id = request.POST.get('entrepot')
        commande.notes = request.POST.get('notes', '')
        commande.save()
        
        # Supprimer les anciennes lignes et créer les nouvelles
        commande.lignecommandevente_set.all().delete()
        
        produits_ids = request.POST.getlist('produit[]')
        quantites = request.POST.getlist('quantite[]')
        prix = request.POST.getlist('prix_unitaire[]')
        
        for i, produit_id in enumerate(produits_ids):
            if produit_id and quantites[i]:
                produit = Produit.objects.get(pk=produit_id)
                LigneCommandeVente.objects.create(
                    commande=commande,
                    produit_id=produit_id,
                    quantite=int(quantites[i]),
                    prix_unitaire=float(prix[i]) if prix[i] else produit.prix_vente,
                    taux_tva=produit.taux_tva,
                    remise=0
                )
        
        commande.calculer_totaux()
        
        messages.success(request, 'Commande modifiée avec succès!')
        return redirect('ventes:details_commande_vente', pk=commande.pk)
    
    # GET - Afficher le formulaire
    # Supprimez les imports conditionnels ici
    clients = Client.objects.filter(est_actif=True).order_by('nom')
    entrepots = Entrepot.objects.filter(est_actif=True).order_by('nom')
    produits = Produit.objects.filter(est_actif=True).order_by('nom')
    lignes = commande.lignecommandevente_set.select_related('produit').all()
    
    contexte = {
        'commande': commande,
        'lignes': lignes,
        'clients': clients,
        'entrepots': entrepots,
        'produits': produits,
        'action': 'Modifier',
    }
    
    return render(request, 'ventes/formulaire_commande.jinja', contexte)


@login_required
@transaction.atomic
def confirmer_commande_vente(request, pk):
    """Vue pour confirmer une commande de vente"""
    commande = get_object_or_404(CommandeVente, pk=pk)
    
    if commande.statut != 'BROUILLON':
        messages.error(request, 'Cette commande ne peut pas être confirmée.')
        return redirect('ventes:details_commande_vente', pk=pk)
    
    if request.method == 'POST':
        # Vérifier le stock disponible pour chaque ligne
        stocks_insuffisants = []
        for ligne in commande.lignecommandevente_set.all():
            stock_disponible = ligne.produit.stock_actuel
            if stock_disponible < ligne.quantite:
                stocks_insuffisants.append({
                    'produit': ligne.produit.nom,
                    'requis': ligne.quantite,
                    'disponible': stock_disponible
                })
        
        if stocks_insuffisants:
            contexte = {
                'commande': commande,
                'stocks_insuffisants': stocks_insuffisants,
            }
            return render(request, 'ventes/erreur_confirmation.jinja', contexte)
        
        # Confirmer la commande
        commande.statut = 'CONFIRME'
        commande.save()
        
        messages.success(request, f'Commande {commande.numero_commande} confirmée!')
        return redirect('ventes:details_commande_vente', pk=pk)
    
    return render(request, 'ventes/confirmer_commande.jinja', {'commande': commande})


@login_required
@transaction.atomic
def expedier_commande_vente(request, pk):
    """Vue pour expédier une commande de vente"""
    commande = get_object_or_404(CommandeVente, pk=pk)
    
    if commande.statut != 'CONFIRME':
        messages.error(request, 'Cette commande doit être confirmée avant expédition.')
        return redirect('ventes:details_commande_vente', pk=pk)
    
    if request.method == 'POST':
        # Créer les mouvements de stock pour chaque ligne
        for ligne in commande.lignecommandevente_set.all():
            MouvementStock.objects.create(
                produit=ligne.produit,
                entrepot=commande.entrepot,
                type_mouvement='SORTIE',
                quantite=-ligne.quantite,  # Négatif pour sortie de stock
                reference=commande.numero_commande,
                notes=f'Expédition commande {commande.numero_commande}',
                utilisateur=request.user
            )
        
        commande.statut = 'EXPEDIE'
        commande.save()
        
        messages.success(request, f'Commande {commande.numero_commande} expédiée!')
        return redirect('ventes:details_commande_vente', pk=pk)
    
    lignes = commande.lignecommandevente_set.select_related('produit').all()
    contexte = {
        'commande': commande, 
        'lignes': lignes,
        'aujourdhui': date.today(),
        }
    return render(request, 'ventes/expedier_commande.jinja', contexte)

# VERSION DEBUG - Remplace ta fonction liste_expeditions par celle-ci temporairement

@login_required
def liste_expeditions(request):
    """
    Liste des commandes expédiées et à expédier - VERSION DEBUG
    """
    from django.db.models import Sum
    
    # Filtres
    statut_filtre = request.GET.get('statut', '')
    search_query = request.GET.get('search', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Base queryset
    expeditions = CommandeVente.objects.filter(
        statut__in=['CONFIRME', 'EXPEDIE', 'LIVRE']
    ).select_related('client')
    
    # DEBUG: Afficher le nombre de commandes
    print(f"DEBUG: Nombre d'expéditions trouvées: {expeditions.count()}")
    
    # DEBUG: Afficher les premières commandes
    for exp in expeditions[:3]:
        print(f"DEBUG: Commande {exp.numero_commande} - Client: {exp.client.nom} - Total: {exp.total}")
    
    # Filtres
    if statut_filtre:
        expeditions = expeditions.filter(statut=statut_filtre)
    
    if search_query:
        expeditions = expeditions.filter(
            Q(numero_commande__icontains=search_query) |
            Q(client__nom__icontains=search_query)
        )
    
    if date_debut:
        expeditions = expeditions.filter(date_commande__gte=date_debut)
    
    if date_fin:
        expeditions = expeditions.filter(date_commande__lte=date_fin)
    
    # Statistiques
    stats = {
        'a_expedier': CommandeVente.objects.filter(statut='CONFIRME').count(),
        'en_cours': CommandeVente.objects.filter(statut='EXPEDIE').count(),
        'livrees': CommandeVente.objects.filter(statut='LIVRE').count(),
        'total': CommandeVente.objects.filter(statut__in=['CONFIRME', 'EXPEDIE', 'LIVRE']).count(),
    }
    
    # Total des expéditions
    result = expeditions.aggregate(total=Sum('total'))
    total_expeditions = result['total'] if result['total'] is not None else Decimal('0')
    
    # DEBUG: Afficher le total
    print(f"DEBUG: Total expéditions: {total_expeditions}")
    
    context = {
        'expeditions': expeditions.order_by('-date_commande'),
        'stats': stats,
        'total_expeditions': total_expeditions,
        'statut_filtre': statut_filtre,
        'search_query': search_query,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    
    return render(request, 'ventes/liste_expeditions.jinja', context)
@login_required
def details_expedition(request, pk):
    """
    Détails d'une expédition
    """
    commande = get_object_or_404(
        CommandeVente.objects.select_related('client').prefetch_related('lignes__produit'),
        pk=pk
    )
    
    context = {
        'commande': commande,
    }
    
    return render(request, 'ventes/details_expedition.jinja', context)


@login_required
def details_expedition(request, pk):
    """
    Détails d'une expédition
    """
    commande = get_object_or_404(
        CommandeVente.objects.select_related('client').prefetch_related('lignes__produit'),
        pk=pk
    )
    
    context = {
        'commande': commande,
    }
    
    return render(request, 'ventes/details_expedition.jinja', context)


@login_required
@transaction.atomic
def facturer_commande_vente(request, pk):
    """Vue pour créer une facture à partir d'une commande de vente"""
    commande = get_object_or_404(CommandeVente, pk=pk)
    
    if commande.statut != 'EXPEDIE':
        messages.error(request, 'Cette commande doit être expédiée avant facturation.')
        return redirect('ventes:details_commande_vente', pk=pk)
    
    # Générer le numéro de facture
    aujourdhui = datetime.now()
    prefixe = f"FAC{aujourdhui.strftime('%Y%m')}"
    derniere_facture = Facture.objects.filter(
        numero_facture__startswith=prefixe
    ).order_by('-numero_facture').first()
    
    if derniere_facture:
        dernier_numero = int(derniere_facture.numero_facture[-4:])
        nouveau_numero = dernier_numero + 1
    else:
        nouveau_numero = 1
    
    numero_facture = f"{prefixe}{nouveau_numero:04d}"
    
    # Créer la facture
    facture = Facture.objects.create(
        numero_facture=numero_facture,
        commande_vente=commande,
        client=commande.client,
        date_echeance=aujourdhui.date() + timedelta(days=30),
        sous_total=commande.sous_total,
        montant_tva=commande.montant_tva,
        total=commande.total,
        statut='BROUILLON'
    )
    
    commande.statut = 'FACTURE'
    commande.save()
    
    messages.success(request, f'Facture {numero_facture} créée!')
    return redirect('ventes:details_facture', pk=facture.pk)

@login_required
@transaction.atomic
def supprimer_commande_vente(request, pk):
    """Vue pour supprimer une commande de vente"""
    commande = get_object_or_404(CommandeVente, pk=pk)
    
    # Vérifier si la commande peut être supprimée
    if commande.statut != 'BROUILLON':
        messages.error(request, 'Seules les commandes au statut "Brouillon" peuvent être supprimées.')
        return redirect('ventes:liste_commandes_vente')
    
    if request.method == 'POST':
        numero = commande.numero_commande
        commande.delete()
        messages.success(request, f'Commande {numero} supprimée avec succès.')
        return redirect('ventes:liste_commandes_vente')
    
    # Si la méthode est GET, afficher une page de confirmation
    return render(request, 'ventes/confirmer_suppression_commande.jinja', {'commande': commande})


# ========== GESTION DES FACTURES ==========

@login_required
def liste_factures(request):
    """Vue pour afficher la liste des factures"""
    filtre_statut = request.GET.get('statut', '')
    recherche = request.GET.get('recherche', '')
    
    factures = Facture.objects.select_related('client', 'commande_vente')
    
    if filtre_statut:
        factures = factures.filter(statut=filtre_statut)
    
    if recherche:
        factures = factures.filter(
            Q(numero_facture__icontains=recherche) |
            Q(client__nom__icontains=recherche)
        )
    
    factures = factures.order_by('-date_creation')
    
    contexte = {
        'factures': factures,
        'filtre_statut': filtre_statut,
        'recherche': recherche,
        'statuts': Facture.STATUTS,
        'aujourdhui': date.today().isoformat(),
    }
    
    return render(request, 'ventes/liste_factures.jinja', contexte)


@login_required
def details_facture(request, pk):
    """Vue pour afficher les détails d'une facture"""
    facture = get_object_or_404(
        Facture.objects.select_related('client', 'commande_vente'),
        pk=pk
    )
    
    # Récupérer les lignes de la facture (avec related_name='lignes')
    lignes = facture.lignes.select_related('produit').all()
    
    # Si la facture n'a pas de lignes mais a une commande associée,
    # récupérer les lignes de la commande (cas de transition)
    if not lignes.exists() and facture.commande_vente:
        lignes = facture.commande_vente.lignecommandevente_set.select_related('produit').all()
    
    contexte = {
        'facture': facture,
        'lignes': lignes,
    }
    
    return render(request, 'ventes/details_facture.jinja', contexte)


@login_required
def envoyer_facture(request, pk):
    """Vue pour envoyer une facture au client (version simplifiée)"""
    facture = get_object_or_404(Facture, pk=pk)
    
    if facture.statut not in ['BROUILLON', 'ENVOYEE']:
        messages.error(request, 'Cette facture ne peut pas être envoyée.')
        return redirect('ventes:details_facture', pk=pk)
    
    if request.method == 'POST':
        facture.statut = 'ENVOYEE'
        facture.save()
        
        # Optionnel: Enregistrer l'action dans l'historique
        try:
            from .models import HistoriqueAction
            HistoriqueAction.objects.create(
                utilisateur=request.user,
                action=f'A marqué la facture {facture.numero_facture} comme envoyée',
                modele='Facture',
                objet_id=facture.pk
            )
        except (ImportError, NameError):
            pass
        
        messages.success(request, f'Facture {facture.numero_facture} marquée comme envoyée avec succès!')
        return redirect('ventes:details_facture', pk=pk)
    
    # GET - Afficher la page de confirmation
    return render(request, 'ventes/envoyer_facture.jinja', {'facture': facture})


@login_required
@transaction.atomic
def enregistrer_paiement(request, pk):
    """Vue pour enregistrer un paiement sur une facture"""
    facture = get_object_or_404(Facture, pk=pk)
    
    if request.method == 'POST':
        # Récupérer les données du formulaire
        montant_str = request.POST.get('montant', '0')
        date_paiement = request.POST.get('date_paiement')
        mode_paiement = request.POST.get('mode_paiement', 'ESPECES')
        reference = request.POST.get('reference', '')
        
        try:
            # Convertir en Decimal pour éviter l'erreur de type
            montant = Decimal(montant_str)
        except (ValueError, InvalidOperation):
            messages.error(request, 'Le montant doit être un nombre valide.')
            return redirect('ventes:details_facture', pk=pk)
        
        if montant <= Decimal('0'):
            messages.error(request, 'Le montant doit être supérieur à 0.')
            return redirect('ventes:details_facture', pk=pk)
        
        # Calculer le solde actuel en Decimal
        total = facture.total if facture.total else Decimal('0')
        paye = facture.montant_paye if facture.montant_paye else Decimal('0')
        solde_actuel = total - paye
        
        if montant > solde_actuel:
            messages.error(request, f'Le montant ne peut pas être supérieur au solde de {solde_actuel:,.0f} FCFA.')
            return redirect('ventes:details_facture', pk=pk)
        
        # Mettre à jour le montant payé
        facture.montant_paye += montant
        
        # Si le montant payé >= total, marquer comme payée
        if facture.montant_paye >= total:
            facture.statut = 'PAYEE'
        
        facture.save()
        
        # Optionnel: Enregistrer les détails du paiement si le modèle Paiement existe
        try:
            # Si vous avez un modèle Paiement, vous pouvez l'utiliser
            from .models import Paiement
            Paiement.objects.create(
                facture=facture,
                montant=montant,
                date_paiement=date_paiement,
                mode_paiement=mode_paiement,
                reference=reference,
                utilisateur=request.user
            )
        except (ImportError, NameError):
            # Si vous n'avez pas de modèle Paiement, continuez sans
            pass
        
        messages.success(request, f'Paiement de {montant:,.0f} FCFA enregistré avec succès!')
        return redirect('ventes:details_facture', pk=pk)
    
    # Si méthode GET, afficher le formulaire
    # Assurez-vous de passer la date du jour pour le formulaire
    contexte = {
        'facture': facture,
        'aujourdhui': date.today().isoformat(),
    }
    
    return render(request, 'ventes/enregistrer_paiement.jinja', contexte)

@login_required
def facture_pdf(request, pk):
    """Vue pour générer le PDF de la facture"""
    facture = get_object_or_404(Facture, pk=pk)
    lignes = facture.lignes.all()  # Utilise related_name='lignes'
    
    # Calculer le solde
    total_paye = facture.montant_paye if facture.montant_paye else 0
    solde = facture.total - total_paye
    
    context = {
        'facture': facture,
        'lignes': lignes,
        'total_paye': total_paye,
        'solde': solde,
    }
    
    return render(request, 'ventes/facture_pdf.jinja', context)


# ========== API AJAX ==========

@login_required
def obtenir_prix_produit(request, pk):
    """API AJAX pour obtenir le prix d'un produit"""
    try:
        produit = Produit.objects.get(pk=pk)
        return JsonResponse({
            'succes': True,
            'prix_vente': float(produit.prix_vente),
            'taux_tva': float(produit.taux_tva),
            'stock_actuel': produit.stock_actuel,
        })
    except Produit.DoesNotExist:
        return JsonResponse({'succes': False, 'erreur': 'Produit non trouvé'})
    


@login_required
def imprimer_commande(request, pk):
    """
    Afficher la page d'impression de la commande (version optimisée pour l'impression)
    """
    commande = get_object_or_404(CommandeVente, pk=pk)
    lignes = commande.lignecommandevente_set.select_related('produit').all()
    
    contexte = {
        'commande': commande,
        'lignes': lignes,
        'aujourdhui': timezone.now(),
    }
    
    return render(request, 'ventes/imprimer_commande.jinja', contexte)


@login_required
def telecharger_commande_pdf(request, pk):
    """
    Télécharger la commande en PDF avec mise en page professionnelle
    """
    commande = get_object_or_404(CommandeVente, pk=pk)
    lignes = commande.lignecommandevente_set.select_related('produit').all()
    
    # Créer le buffer pour le PDF
    buffer = BytesIO()
    
    # Créer le document PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=2*cm,
        bottomMargin=2*cm,
        leftMargin=2*cm,
        rightMargin=2*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # ========== EN-TÊTE ==========
    # Style pour le titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#6366f1'),
        spaceAfter=30,
        alignment=1,  # Centré
    )
    
    title = Paragraph(f"Commande {commande.numero_commande}", title_style)
    elements.append(title)
    
    # Date d'impression
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=1,
    )
    date_text = Paragraph(f"Imprimé le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style)
    elements.append(date_text)
    elements.append(Spacer(1, 30))
    
    # ========== INFORMATIONS COMMANDE ==========
    info_data = [
        ['Client:', commande.client.nom],
        ['Code client:', commande.client.code or '-'],
        ['Entrepôt:', commande.entrepot.nom],
        ['Date commande:', commande.date_commande.strftime('%d/%m/%Y')],
        ['Date livraison:', commande.date_livraison.strftime('%d/%m/%Y') if commande.date_livraison else '-'],
        ['Statut:', commande.get_statut_display()],
    ]
    
    info_table = Table(info_data, colWidths=[4*cm, 12*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748b')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1e293b')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 30))
    
    # ========== LIGNES DE COMMANDE ==========
    # Titre section
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=15,
    )
    elements.append(Paragraph("Lignes de commande", section_style))
    
    # Données du tableau
    table_data = [['Produit', 'Qté', 'Prix unit.', 'Remise', 'TVA', 'Total']]
    
    for ligne in lignes:
        table_data.append([
            ligne.produit.nom if ligne.produit else 'Produit supprimé',
            str(ligne.quantite),
            f"{ligne.prix_unitaire:,.0f} FCFA",
            f"{ligne.remise}%" if ligne.remise else "0%",
            f"{ligne.taux_tva}%",
            f"{ligne.total:,.0f} FCFA",
        ])
    
    # Ajouter les totaux
    table_data.append(['', '', '', '', 'Sous-total:', f"{commande.sous_total:,.0f} FCFA"])
    table_data.append(['', '', '', '', 'TVA:', f"{commande.montant_tva:,.0f} FCFA"])
    table_data.append(['', '', '', '', 'TOTAL:', f"{commande.total:,.0f} FCFA"])
    
    # Créer le tableau
    table = Table(table_data, colWidths=[6*cm, 1.5*cm, 2.5*cm, 1.5*cm, 2*cm, 3*cm])
    table.setStyle(TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Contenu
        ('BACKGROUND', (0, 1), (-1, -4), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -4), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -4), 0.5, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        
        # Lignes de totaux
        ('BACKGROUND', (0, -3), (-1, -2), colors.HexColor('#f8fafc')),
        ('FONTNAME', (4, -3), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (4, -3), (-1, -1), 'RIGHT'),
        ('ALIGN', (5, -3), (-1, -1), 'RIGHT'),
        
        # Ligne total final
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0e7ff')),
        ('TEXTCOLOR', (4, -1), (-1, -1), colors.HexColor('#6366f1')),
        ('FONTSIZE', (4, -1), (-1, -1), 11),
        
        # Lignes alternées
        ('ROWBACKGROUNDS', (0, 1), (-1, -4), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    
    elements.append(table)
    
    # ========== NOTES ==========
    if commande.notes:
        elements.append(Spacer(1, 30))
        elements.append(Paragraph("Notes", section_style))
        
        notes_style = ParagraphStyle(
            'NotesStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#78350f'),
            leftIndent=10,
            rightIndent=10,
        )
        notes_text = Paragraph(commande.notes, notes_style)
        elements.append(notes_text)
    
    # ========== FOOTER ==========
    elements.append(Spacer(1, 40))
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1,
    )
    footer_text = Paragraph(
        f"Document généré automatiquement - {getattr(settings, 'SITE_NAME', 'ERP MEA')}",
        footer_style
    )
    elements.append(footer_text)
    
    # Construire le PDF
    doc.build(elements)
    
    # Récupérer le PDF
    pdf = buffer.getvalue()
    buffer.close()
    
    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="commande_{commande.numero_commande}.pdf"'
    response.write(pdf)
    
    return response


@login_required
def envoyer_commande_email(request, pk):
    """
    Envoyer la commande par email au client
    """
    commande = get_object_or_404(CommandeVente, pk=pk)
    
    # Vérifier que le client a un email
    if not commande.client.email:
        messages.error(request, f'Le client {commande.client.nom} n\'a pas d\'adresse email enregistrée.')
        return redirect('ventes:details_commande_vente', pk=pk)
    
    try:
        # Générer le PDF de la commande
        lignes = commande.lignecommandevente_set.select_related('produit').all()
        
        # Créer le buffer pour le PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Titre
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, 
                                     textColor=colors.HexColor('#6366f1'), spaceAfter=30, alignment=1)
        elements.append(Paragraph(f"Commande {commande.numero_commande}", title_style))
        elements.append(Spacer(1, 20))
        
        # Informations
        info_data = [
            ['Client:', commande.client.nom],
            ['Date commande:', commande.date_commande.strftime('%d/%m/%Y')],
            ['Date livraison:', commande.date_livraison.strftime('%d/%m/%Y') if commande.date_livraison else '-'],
            ['Statut:', commande.get_statut_display()],
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 12*cm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 20))
        
        # Lignes de commande
        table_data = [['Produit', 'Quantité', 'Prix unitaire', 'Total']]
        for ligne in lignes:
            table_data.append([
                ligne.produit.nom if ligne.produit else '-',
                str(ligne.quantite),
                f"{ligne.prix_unitaire:,.0f} FCFA",
                f"{ligne.total:,.0f} FCFA",
            ])
        
        table_data.append(['', '', 'TOTAL:', f"{commande.total:,.0f} FCFA"])
        
        table = Table(table_data, colWidths=[7*cm, 2*cm, 3.5*cm, 3.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#e2e8f0')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0e7ff')),
            ('FONTNAME', (2, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (2, -1), (-1, -1), 'RIGHT'),
        ]))
        elements.append(table)
        
        if commande.notes:
            elements.append(Spacer(1, 20))
            elements.append(Paragraph(f"<b>Notes:</b> {commande.notes}", styles['Normal']))
        
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        
        # Préparer l'email
        subject = f'Commande {commande.numero_commande} - {commande.client.nom}'
        
        # Corps de l'email en HTML
        html_message = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <h2 style="color: #6366f1; border-bottom: 2px solid #6366f1; padding-bottom: 10px;">
                    Commande {commande.numero_commande}
                </h2>
                
                <p>Bonjour {commande.client.nom},</p>
                
                <p>Veuillez trouver ci-joint les détails de votre commande.</p>
                
                <div style="background: #f8fafc; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <table style="width: 100%;">
                        <tr>
                            <td style="padding: 5px;"><strong>Numéro de commande:</strong></td>
                            <td style="padding: 5px;">{commande.numero_commande}</td>
                        </tr>
                        <tr>
                            <td style="padding: 5px;"><strong>Date:</strong></td>
                            <td style="padding: 5px;">{commande.date_commande.strftime('%d/%m/%Y')}</td>
                        </tr>
                        <tr>
                            <td style="padding: 5px;"><strong>Montant total:</strong></td>
                            <td style="padding: 5px;"><strong style="color: #6366f1;">{commande.total:,.0f} FCFA</strong></td>
                        </tr>
                        <tr>
                            <td style="padding: 5px;"><strong>Statut:</strong></td>
                            <td style="padding: 5px;">{commande.get_statut_display()}</td>
                        </tr>
                    </table>
                </div>
                
                <p>Pour toute question, n'hésitez pas à nous contacter.</p>
                
                <p>Cordialement,<br>
                <strong>{getattr(settings, 'SITE_NAME', 'ERP MEA')}</strong></p>
                
                <hr style="margin: 30px 0; border: none; border-top: 1px solid #e2e8f0;">
                <p style="font-size: 12px; color: #64748b;">
                    Cet email a été généré automatiquement. Merci de ne pas y répondre.
                </p>
            </div>
        </body>
        </html>
        """
        
        # Créer l'email
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@erp-mea.com'),
            to=[commande.client.email],
        )
        email.content_subtype = 'html'  # Email en HTML
        
        # Attacher le PDF
        email.attach(f'commande_{commande.numero_commande}.pdf', pdf, 'application/pdf')
        
        # Envoyer l'email
        email.send()
        
        messages.success(request, f'La commande a été envoyée par email à {commande.client.email}')
        
    except Exception as e:
        messages.error(request, f'Erreur lors de l\'envoi de l\'email: {str(e)}')
    
    return redirect('ventes:details_commande_vente', pk=pk)

# ========== ACTIONS RAPIDES SUR LES FACTURES ==========
# À ajouter dans ventes/views.py

@login_required
def imprimer_facture(request, pk):
    """
    Afficher la page d'impression de la facture (version optimisée pour l'impression)
    """
    facture = get_object_or_404(Facture, pk=pk)
    lignes = facture.lignes.select_related('produit').all()
    
    # Calculer le solde
    total = facture.total if facture.total else 0
    paye = facture.montant_paye if facture.montant_paye else 0
    solde = total - paye
    
    contexte = {
        'facture': facture,
        'lignes': lignes,
        'aujourdhui': datetime.now(),
        'solde': solde,
    }
    
    return render(request, 'ventes/imprimer_facture.jinja', contexte)


@login_required
def telecharger_facture_pdf(request, pk):
    """
    Télécharger la facture en PDF - VERSION CORRIGÉE SANS CHAMPS MANQUANTS
    """
    facture = get_object_or_404(Facture, pk=pk)
    lignes = facture.lignes.select_related('produit').all()
    
    # Calculer le solde
    total = facture.total if facture.total else 0
    paye = facture.montant_paye if facture.montant_paye else 0
    solde = total - paye
    
    # Calculer le taux de TVA (18% par défaut)
    taux_tva = 18
    if facture.montant_tva and facture.sous_total and facture.sous_total > 0:
       taux_tva = round((facture.montant_tva / facture.sous_total) * 100, 2)

   

    # Créer le buffer pour le PDF
    buffer = BytesIO()
    
    # Créer le document PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=2*cm,
        bottomMargin=2*cm,
        leftMargin=2*cm,
        rightMargin=2*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # ========== EN-TÊTE ==========
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.HexColor('#6366f1'),
        spaceAfter=30,
        alignment=1,
    )
    
    title = Paragraph(f"FACTURE {facture.numero_facture}", title_style)
    elements.append(title)
    
    # Date d'impression
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=1,
    )
    date_text = Paragraph(f"Imprimé le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style)
    elements.append(date_text)
    elements.append(Spacer(1, 30))
    
    # ========== INFORMATIONS FACTURE ==========
    info_data = [
        ['Client:', facture.client.nom],
        ['Code client:', facture.client.code or '-'],
        ['Date facture:', facture.date_facture.strftime('%d/%m/%Y') if facture.date_facture else 
                         facture.date_creation.strftime('%d/%m/%Y') if facture.date_creation else '-'],
        ['Date échéance:', facture.date_echeance.strftime('%d/%m/%Y') if facture.date_echeance else '-'],
        ['Statut:', facture.get_statut_display()],
    ]
    
    info_table = Table(info_data, colWidths=[4*cm, 12*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748b')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1e293b')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 30))
    
    # ========== LIGNES DE FACTURE ==========
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=15,
    )
    elements.append(Paragraph("Articles facturés", section_style))
    
    # Données du tableau
    table_data = [['Description', 'Qté', 'Prix unit.', 'Total']]
    
    for ligne in lignes:
        table_data.append([
            ligne.produit.nom if ligne.produit else 'Produit supprimé',
            str(ligne.quantite),
            f"{ligne.prix_unitaire:,.0f} FCFA",
            f"{(ligne.prix_unitaire * ligne.quantite):,.0f} FCFA",
        ])
    
    # Ajouter les totaux
    table_data.append(['', '', 'Sous-total HT:', f"{facture.sous_total:,.0f} FCFA" if facture.sous_total else '0 FCFA'])
    if facture.montant_tva and facture.montant_tva > 0:
        table_data.append(['', '', f'TVA ({taux_tva}%):', f"{facture.montant_tva:,.0f} FCFA"])
    table_data.append(['', '', 'TOTAL TTC:', f"{total:,.0f} FCFA"])
    table_data.append(['', '', 'Montant payé:', f"{paye:,.0f} FCFA"])
    table_data.append(['', '', 'SOLDE:', f"{solde:,.0f} FCFA"])
    
    # Créer le tableau
    table = Table(table_data, colWidths=[9*cm, 2*cm, 2.5*cm, 3*cm])
    table.setStyle(TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Contenu
        ('BACKGROUND', (0, 1), (-1, -6), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -6), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -6), 0.5, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        
        # Lignes de totaux
        ('BACKGROUND', (0, -5), (-1, -2), colors.HexColor('#f8fafc')),
        ('FONTNAME', (2, -5), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (2, -5), (-1, -1), 'RIGHT'),
        ('ALIGN', (3, -5), (-1, -1), 'RIGHT'),
        
        # Ligne total TTC
        ('BACKGROUND', (0, -3), (-1, -3), colors.HexColor('#e0e7ff')),
        ('TEXTCOLOR', (2, -3), (-1, -3), colors.HexColor('#6366f1')),
        ('FONTSIZE', (2, -3), (-1, -3), 11),
        
        # Ligne solde final
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fef3c7') if solde > 0 else colors.HexColor('#dcfce7')),
        ('TEXTCOLOR', (2, -1), (-1, -1), colors.HexColor('#f59e0b') if solde > 0 else colors.HexColor('#10b981')),
        ('FONTSIZE', (2, -1), (-1, -1), 12),
        
        # Lignes alternées
        ('ROWBACKGROUNDS', (0, 1), (-1, -6), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    
    elements.append(table)
    
    # ========== REMARQUES (SUPPRIMÉ - CHAMP N'EXISTE PAS) ==========
    # Anciennement: if facture.remarques: ...
    # Maintenant: On ne fait rien car le champ n'existe pas
    
    # ========== FOOTER ==========
    elements.append(Spacer(1, 40))
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1,
    )
    footer_text = Paragraph(
        f"Document généré automatiquement - {getattr(settings, 'SITE_NAME', 'ERP MEA')}",
        footer_style
    )
    elements.append(footer_text)
    
    # Construire le PDF
    doc.build(elements)
    
    # Récupérer le PDF
    pdf = buffer.getvalue()
    buffer.close()
    
    # Créer la réponse HTTP avec téléchargement direct
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="facture_{facture.numero_facture}.pdf"'
    response.write(pdf)
    
    return response


@login_required
def envoyer_facture_email(request, pk):
    """
    Envoyer la facture par email - VERSION FINALE SANS CHAMPS MANQUANTS
    """
    facture = get_object_or_404(Facture, pk=pk)
    
    if request.method == 'POST':
        # Vérifier que le client a un email
        if not facture.client.email:
            messages.warning(request, f'Le client {facture.client.nom} n\'a pas d\'adresse email. La facture a été marquée comme "Envoyée".')
            facture.statut = 'ENVOYEE'
            facture.save()
            return redirect('ventes:details_facture', pk=pk)
        
        try:
            # Générer le PDF de la facture
            lignes = facture.lignes.select_related('produit').all()
            total = facture.total if facture.total else 0
            paye = facture.montant_paye if facture.montant_paye else 0
            solde = total - paye
            
            # Calculer le taux de TVA
            taux_tva = 18
            if facture.montant_tva and facture.sous_total and facture.sous_total > 0:
                taux_tva = round((facture.montant_tva / facture.sous_total) * 100, 2)
            
            # Créer le buffer pour le PDF
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Titre
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=28, 
                                         textColor=colors.HexColor('#6366f1'), spaceAfter=30, alignment=1)
            elements.append(Paragraph(f"FACTURE {facture.numero_facture}", title_style))
            elements.append(Spacer(1, 20))
            
            # Informations
            info_data = [
                ['Client:', facture.client.nom],
                ['Date facture:', facture.date_facture.strftime('%d/%m/%Y') if facture.date_facture else 
                                 facture.date_creation.strftime('%d/%m/%Y') if facture.date_creation else '-'],
                ['Date échéance:', facture.date_echeance.strftime('%d/%m/%Y') if facture.date_echeance else '-'],
                ['Statut:', facture.get_statut_display()],
            ]
            
            info_table = Table(info_data, colWidths=[4*cm, 12*cm])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 20))
            
            # Lignes de facture
            table_data = [['Description', 'Quantité', 'Prix unitaire', 'Total']]
            for ligne in lignes:
                table_data.append([
                    ligne.produit.nom if ligne.produit else '-',
                    str(ligne.quantite),
                    f"{ligne.prix_unitaire:,.0f} FCFA",
                    f"{(ligne.prix_unitaire * ligne.quantite):,.0f} FCFA",
                ])
            
            # Totaux
            table_data.append(['', '', 'Sous-total HT:', f"{facture.sous_total:,.0f} FCFA" if facture.sous_total else '0 FCFA'])
            if facture.montant_tva and facture.montant_tva > 0:
                table_data.append(['', '', f'TVA ({taux_tva}%):', f"{facture.montant_tva:,.0f} FCFA"])
            table_data.append(['', '', 'TOTAL TTC:', f"{total:,.0f} FCFA"])
            table_data.append(['', '', 'Montant payé:', f"{paye:,.0f} FCFA"])
            table_data.append(['', '', 'SOLDE:', f"{solde:,.0f} FCFA"])
            
            table = Table(table_data, colWidths=[8*cm, 2*cm, 3*cm, 3*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -6), 0.5, colors.HexColor('#e2e8f0')),
                ('BACKGROUND', (0, -5), (-1, -2), colors.HexColor('#f8fafc')),
                ('BACKGROUND', (0, -3), (-1, -3), colors.HexColor('#e0e7ff')),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fef3c7') if solde > 0 else colors.HexColor('#dcfce7')),
                ('FONTNAME', (2, -5), (-1, -1), 'Helvetica-Bold'),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('ALIGN', (2, -5), (-1, -1), 'RIGHT'),
                ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(table)
            
            # REMARQUES SUPPRIMÉES - Le champ n'existe pas
            # if facture.remarques: ...
            
            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()
            
            # Préparer l'email
            subject = f'Facture {facture.numero_facture} - {facture.client.nom}'
            
            # Corps de l'email en HTML
            html_message = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h2 style="color: #6366f1; border-bottom: 2px solid #6366f1; padding-bottom: 10px;">
                        Facture {facture.numero_facture}
                    </h2>
                    
                    <p>Bonjour {facture.client.nom},</p>
                    
                    <p>Veuillez trouver ci-joint votre facture.</p>
                    
                    <div style="background: #f8fafc; padding: 15px; border-radius: 8px; margin: 20px 0;">
                        <table style="width: 100%;">
                            <tr>
                                <td style="padding: 5px;"><strong>Numéro de facture:</strong></td>
                                <td style="padding: 5px;">{facture.numero_facture}</td>
                            </tr>
                            <tr>
                                <td style="padding: 5px;"><strong>Date:</strong></td>
                                <td style="padding: 5px;">{facture.date_facture.strftime('%d/%m/%Y') if facture.date_facture else facture.date_creation.strftime('%d/%m/%Y') if facture.date_creation else '-'}</td>
                            </tr>
                            <tr>
                                <td style="padding: 5px;"><strong>Montant total:</strong></td>
                                <td style="padding: 5px;"><strong style="color: #6366f1;">{total:,.0f} FCFA</strong></td>
                            </tr>
                            <tr>
                                <td style="padding: 5px;"><strong>Solde restant:</strong></td>
                                <td style="padding: 5px;"><strong style="color: {'#f59e0b' if solde > 0 else '#10b981'};">{solde:,.0f} FCFA</strong></td>
                            </tr>
                        </table>
                    </div>
                    
                    <p>Pour toute question, n'hésitez pas à nous contacter.</p>
                    
                    <p>Cordialement,<br>
                    <strong>{getattr(settings, 'SITE_NAME', 'ERP MEA')}</strong></p>
                    
                    <hr style="margin: 30px 0; border: none; border-top: 1px solid #e2e8f0;">
                    <p style="font-size: 12px; color: #64748b;">
                        Cet email a été généré automatiquement. Merci de ne pas y répondre.
                    </p>
                </div>
            </body>
            </html>
            """
            
            # Créer l'email
            email = EmailMessage(
                subject=subject,
                body=html_message,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@erp-mea.com'),
                to=[facture.client.email],
            )
            email.content_subtype = 'html'
            
            # Attacher le PDF
            email.attach(f'facture_{facture.numero_facture}.pdf', pdf, 'application/pdf')
            
            # Envoyer l'email
            email.send()
            
            # Mettre à jour le statut
            facture.statut = 'ENVOYEE'
            facture.save()
            
            messages.success(request, f'La facture a été envoyée par email à {facture.client.email}')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'envoi de l\'email: {str(e)}')
        
        return redirect('ventes:details_facture', pk=pk)
    
    # GET - Page de confirmation
    contexte = {
        'facture': facture,
    }
    return render(request, 'ventes/envoyer_facture.jinja', contexte)
    """
    Envoyer la facture par email au client
    """
    facture = get_object_or_404(Facture, pk=pk)
    
    if request.method == 'POST':
        # Vérifier que le client a un email
        if not facture.client.email:
            messages.warning(request, f'Le client {facture.client.nom} n\'a pas d\'adresse email. La facture a été marquée comme "Envoyée".')
            facture.statut = 'ENVOYEE'
            facture.save()
            return redirect('ventes:details_facture', pk=pk)
        
        try:
            # Générer le PDF de la facture
            lignes = facture.lignes.select_related('produit').all()
            total = facture.total if facture.total else 0
            paye = facture.montant_paye if facture.montant_paye else 0
            solde = total - paye
            
            # Calculer le taux de TVA
            taux_tva = 18  # Taux par défaut
            if facture.montant_tva and facture.sous_total and facture.sous_total > 0:
                taux_tva = round((facture.montant_tva / facture.sous_total) * 100, 2)
            
            # Créer le buffer pour le PDF
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Titre
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=28, 
                                         textColor=colors.HexColor('#6366f1'), spaceAfter=30, alignment=1)
            elements.append(Paragraph(f"FACTURE {facture.numero_facture}", title_style))
            elements.append(Spacer(1, 20))
            
            # Informations
            info_data = [
                ['Client:', facture.client.nom],
                ['Date facture:', facture.date_facture.strftime('%d/%m/%Y') if facture.date_facture else 
                                 facture.date_creation.strftime('%d/%m/%Y') if facture.date_creation else '-'],
                ['Date échéance:', facture.date_echeance.strftime('%d/%m/%Y') if facture.date_echeance else '-'],
                ['Statut:', facture.get_statut_display()],
            ]
            
            info_table = Table(info_data, colWidths=[4*cm, 12*cm])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 20))
            
            # Lignes de facture
            table_data = [['Description', 'Quantité', 'Prix unitaire', 'Total']]
            for ligne in lignes:
                table_data.append([
                    ligne.produit.nom if ligne.produit else '-',
                    str(ligne.quantite),
                    f"{ligne.prix_unitaire:,.0f} FCFA",
                    f"{(ligne.prix_unitaire * ligne.quantite):,.0f} FCFA",
                ])
            
            # Totaux
            table_data.append(['', '', 'Sous-total HT:', f"{facture.sous_total:,.0f} FCFA" if facture.sous_total else '0 FCFA'])
            if facture.montant_tva and facture.montant_tva > 0:
                table_data.append(['', '', f'TVA ({taux_tva}%):', f"{facture.montant_tva:,.0f} FCFA"])
            table_data.append(['', '', 'TOTAL TTC:', f"{total:,.0f} FCFA"])
            table_data.append(['', '', 'Montant payé:', f"{paye:,.0f} FCFA"])
            table_data.append(['', '', 'SOLDE:', f"{solde:,.0f} FCFA"])
            
            table = Table(table_data, colWidths=[8*cm, 2*cm, 3*cm, 3*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -6), 0.5, colors.HexColor('#e2e8f0')),
                ('BACKGROUND', (0, -5), (-1, -2), colors.HexColor('#f8fafc')),
                ('BACKGROUND', (0, -3), (-1, -3), colors.HexColor('#e0e7ff')),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fef3c7') if solde > 0 else colors.HexColor('#dcfce7')),
                ('FONTNAME', (2, -5), (-1, -1), 'Helvetica-Bold'),
                ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
                ('ALIGN', (2, -5), (-1, -1), 'RIGHT'),
                ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(table)
            
            if facture.remarques:
                elements.append(Spacer(1, 20))
                elements.append(Paragraph(f"<b>Remarques:</b> {facture.remarques}", styles['Normal']))
            
            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()
            
            # Préparer l'email
            subject = f'Facture {facture.numero_facture} - {facture.client.nom}'
            
            # Corps de l'email en HTML
            html_message = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h2 style="color: #6366f1; border-bottom: 2px solid #6366f1; padding-bottom: 10px;">
                        Facture {facture.numero_facture}
                    </h2>
                    
                    <p>Bonjour {facture.client.nom},</p>
                    
                    <p>Veuillez trouver ci-joint votre facture.</p>
                    
                    <div style="background: #f8fafc; padding: 15px; border-radius: 8px; margin: 20px 0;">
                        <table style="width: 100%;">
                            <tr>
                                <td style="padding: 5px;"><strong>Numéro de facture:</strong></td>
                                <td style="padding: 5px;">{facture.numero_facture}</td>
                            </tr>
                            <tr>
                                <td style="padding: 5px;"><strong>Date:</strong></td>
                                <td style="padding: 5px;">{facture.date_facture.strftime('%d/%m/%Y') if facture.date_facture else facture.date_creation.strftime('%d/%m/%Y') if facture.date_creation else '-'}</td>
                            </tr>
                            <tr>
                                <td style="padding: 5px;"><strong>Montant total:</strong></td>
                                <td style="padding: 5px;"><strong style="color: #6366f1;">{total:,.0f} FCFA</strong></td>
                            </tr>
                            <tr>
                                <td style="padding: 5px;"><strong>Solde restant:</strong></td>
                                <td style="padding: 5px;"><strong style="color: {'#f59e0b' if solde > 0 else '#10b981'};">{solde:,.0f} FCFA</strong></td>
                            </tr>
                        </table>
                    </div>
                    
                    <p>Pour toute question, n'hésitez pas à nous contacter.</p>
                    
                    <p>Cordialement,<br>
                    <strong>{getattr(settings, 'SITE_NAME', 'ERP MEA')}</strong></p>
                    
                    <hr style="margin: 30px 0; border: none; border-top: 1px solid #e2e8f0;">
                    <p style="font-size: 12px; color: #64748b;">
                        Cet email a été généré automatiquement. Merci de ne pas y répondre.
                    </p>
                </div>
            </body>
            </html>
            """
            
            # Créer l'email
            email = EmailMessage(
                subject=subject,
                body=html_message,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@erp-mea.com'),
                to=[facture.client.email],
            )
            email.content_subtype = 'html'  # Email en HTML
            
            # Attacher le PDF
            email.attach(f'facture_{facture.numero_facture}.pdf', pdf, 'application/pdf')
            
            # Envoyer l'email
            email.send()
            
            # Mettre à jour le statut de la facture
            facture.statut = 'ENVOYEE'
            facture.save()
            
            messages.success(request, f'La facture a été envoyée par email à {facture.client.email}')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'envoi de l\'email: {str(e)}')
        
        return redirect('ventes:details_facture', pk=pk)
    
    # GET - Afficher la page de confirmation
    contexte = {
        'facture': facture,
    }
    return render(request, 'ventes/envoyer_facture.jinja', contexte)

    """
    Envoyer la facture par email au client
    """
    facture = get_object_or_404(Facture, pk=pk)
    
    if request.method == 'POST':
        # Vérifier que le client a un email
        if not facture.client.email:
            messages.warning(request, f'Le client {facture.client.nom} n\'a pas d\'adresse email. La facture a été marquée comme "Envoyée".')
            facture.statut = 'ENVOYEE'
            facture.save()
            return redirect('ventes:details_facture', pk=pk)
        
        try:
            # Générer le PDF de la facture
            lignes = facture.lignes.select_related('produit').all()
            total = facture.total if facture.total else 0
            paye = facture.montant_paye if facture.montant_paye else 0
            solde = total - paye
            
            # Créer le buffer pour le PDF
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Titre
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=28, 
                                         textColor=colors.HexColor('#6366f1'), spaceAfter=30, alignment=1)
            elements.append(Paragraph(f"FACTURE {facture.numero_facture}", title_style))
            elements.append(Spacer(1, 20))
            
            # Informations
            info_data = [
                ['Client:', facture.client.nom],
                ['Date facture:', facture.date_facture.strftime('%d/%m/%Y') if facture.date_facture else 
                                 facture.date_creation.strftime('%d/%m/%Y') if facture.date_creation else '-'],
                ['Date échéance:', facture.date_echeance.strftime('%d/%m/%Y') if facture.date_echeance else '-'],
                ['Statut:', facture.get_statut_display()],
            ]
            
            info_table = Table(info_data, colWidths=[4*cm, 12*cm])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 20))
            
            # Lignes de facture
            table_data = [['Description', 'Quantité', 'Prix unitaire', 'Total']]
            for ligne in lignes:
                table_data.append([
                    ligne.produit.nom if ligne.produit else '-',
                    str(ligne.quantite),
                    f"{ligne.prix_unitaire:,.0f} FCFA",
                    f"{(ligne.prix_unitaire * ligne.quantite):,.0f} FCFA",
                ])
            
            table_data.append(['', '', 'TOTAL TTC:', f"{total:,.0f} FCFA"])
            table_data.append(['', '', 'Montant payé:', f"{paye:,.0f} FCFA"])
            table_data.append(['', '', 'SOLDE:', f"{solde:,.0f} FCFA"])
            
            table = Table(table_data, colWidths=[8*cm, 2*cm, 3*cm, 3*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -4), 0.5, colors.HexColor('#e2e8f0')),
                ('BACKGROUND', (0, -3), (-1, -3), colors.HexColor('#e0e7ff')),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fef3c7') if solde > 0 else colors.HexColor('#dcfce7')),
                ('FONTNAME', (2, -3), (-1, -1), 'Helvetica-Bold'),
                ('ALIGN', (2, -3), (-1, -1), 'RIGHT'),
            ]))
            elements.append(table)
            
            if facture.remarques:
                elements.append(Spacer(1, 20))
                elements.append(Paragraph(f"<b>Remarques:</b> {facture.remarques}", styles['Normal']))
            
            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()
            
            # Préparer l'email
            subject = f'Facture {facture.numero_facture} - {facture.client.nom}'
            
            # Corps de l'email en HTML
            html_message = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h2 style="color: #6366f1; border-bottom: 2px solid #6366f1; padding-bottom: 10px;">
                        Facture {facture.numero_facture}
                    </h2>
                    
                    <p>Bonjour {facture.client.nom},</p>
                    
                    <p>Veuillez trouver ci-joint votre facture.</p>
                    
                    <div style="background: #f8fafc; padding: 15px; border-radius: 8px; margin: 20px 0;">
                        <table style="width: 100%;">
                            <tr>
                                <td style="padding: 5px;"><strong>Numéro de facture:</strong></td>
                                <td style="padding: 5px;">{facture.numero_facture}</td>
                            </tr>
                            <tr>
                                <td style="padding: 5px;"><strong>Date:</strong></td>
                                <td style="padding: 5px;">{facture.date_facture.strftime('%d/%m/%Y') if facture.date_facture else facture.date_creation.strftime('%d/%m/%Y') if facture.date_creation else '-'}</td>
                            </tr>
                            <tr>
                                <td style="padding: 5px;"><strong>Montant total:</strong></td>
                                <td style="padding: 5px;"><strong style="color: #6366f1;">{total:,.0f} FCFA</strong></td>
                            </tr>
                            <tr>
                                <td style="padding: 5px;"><strong>Solde restant:</strong></td>
                                <td style="padding: 5px;"><strong style="color: {'#f59e0b' if solde > 0 else '#10b981'};">{solde:,.0f} FCFA</strong></td>
                            </tr>
                        </table>
                    </div>
                    
                    <p>Pour toute question, n'hésitez pas à nous contacter.</p>
                    
                    <p>Cordialement,<br>
                    <strong>{getattr(settings, 'SITE_NAME', 'ERP MEA')}</strong></p>
                    
                    <hr style="margin: 30px 0; border: none; border-top: 1px solid #e2e8f0;">
                    <p style="font-size: 12px; color: #64748b;">
                        Cet email a été généré automatiquement. Merci de ne pas y répondre.
                    </p>
                </div>
            </body>
            </html>
            """
            
            # Créer l'email
            email = EmailMessage(
                subject=subject,
                body=html_message,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@erp-mea.com'),
                to=[facture.client.email],
            )
            email.content_subtype = 'html'  # Email en HTML
            
            # Attacher le PDF
            email.attach(f'facture_{facture.numero_facture}.pdf', pdf, 'application/pdf')
            
            # Envoyer l'email
            email.send()
            
            # Mettre à jour le statut de la facture
            facture.statut = 'ENVOYEE'
            facture.save()
            
            messages.success(request, f'La facture a été envoyée par email à {facture.client.email}')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'envoi de l\'email: {str(e)}')
        
        return redirect('ventes:details_facture', pk=pk)
    
    # GET - Afficher la page de confirmation
    contexte = {
        'facture': facture,
    }
    return render(request, 'ventes/envoyer_facture.jinja', contexte)

# ========== FONCTION D'EXPORTATION POUR COMMANDES DE VENTE ==========

import csv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

@login_required
def exporter_commandes_vente(request):
    """
    Vue pour exporter les commandes de vente en Excel, CSV ou PDF
    """
    format_export = request.GET.get('format', 'excel')
    filtre_statut = request.GET.get('statut', '')
    recherche = request.GET.get('recherche', '')
    ids = request.GET.get('ids', '')
    
    # Filtrer les commandes
    commandes = CommandeVente.objects.select_related('client', 'entrepot', 'cree_par')
    
    # Si des IDs spécifiques sont fournis (sélection)
    if ids:
        liste_ids = [int(id) for id in ids.split(',') if id.isdigit()]
        commandes = commandes.filter(id__in=liste_ids)
    else:
        # Sinon appliquer les filtres normaux
        if filtre_statut:
            commandes = commandes.filter(statut=filtre_statut)
        
        if recherche:
            commandes = commandes.filter(
                Q(numero_commande__icontains=recherche) |
                Q(client__nom__icontains=recherche)
            )
    
    commandes = commandes.order_by('-date_creation')
    
    # Appeler la fonction appropriée selon le format
    if format_export == 'excel':
        return exporter_commandes_excel(commandes)
    elif format_export == 'csv':
        return exporter_commandes_csv(commandes)
    elif format_export == 'pdf':
        return exporter_commandes_pdf(commandes)
    else:
        return HttpResponse("Format non supporté", status=400)


def exporter_commandes_excel(commandes):
    """Exporter les commandes en format Excel"""
    from datetime import datetime
    
    # Créer un workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes de Vente"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = [
        'N° Commande',
        'Client',
        'Date Commande',
        'Date Livraison',
        'Statut',
        'Entrepôt',
        'Sous-total',
        'TVA',
        'Total',
        'Créé par'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Données
    for row_idx, commande in enumerate(commandes, start=2):
        data_row = [
            commande.numero_commande,
            commande.client.nom,
            commande.date_commande.strftime('%d/%m/%Y'),
            commande.date_livraison.strftime('%d/%m/%Y') if commande.date_livraison else '',
            commande.get_statut_display(),
            commande.entrepot.nom if commande.entrepot else '',
            float(commande.sous_total),
            float(commande.montant_tva),
            float(commande.total),
            commande.cree_par.username if commande.cree_par else ''
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            # Alignement pour les montants
            if col_idx in [7, 8, 9]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 30, 15, 15, 15, 20, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Ligne de total
    total_row = len(commandes) + 2
    ws.cell(row=total_row, column=6).value = "TOTAL:"
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    ws.cell(row=total_row, column=7).value = sum(float(c.sous_total) for c in commandes)
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = '#,##0.00'
    ws.cell(row=total_row, column=8).value = sum(float(c.montant_tva) for c in commandes)
    ws.cell(row=total_row, column=8).font = Font(bold=True)
    ws.cell(row=total_row, column=8).number_format = '#,##0.00'
    ws.cell(row=total_row, column=9).value = sum(float(c.total) for c in commandes)
    ws.cell(row=total_row, column=9).font = Font(bold=True)
    ws.cell(row=total_row, column=9).number_format = '#,##0.00'
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Préparer la réponse
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"commandes_vente_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def exporter_commandes_csv(commandes):
    """Exporter les commandes en format CSV"""
    from datetime import datetime
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"commandes_vente_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Ajouter BOM UTF-8 pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # En-têtes
    writer.writerow([
        'N° Commande',
        'Client',
        'Date Commande',
        'Date Livraison',
        'Statut',
        'Entrepôt',
        'Sous-total',
        'TVA',
        'Total',
        'Créé par'
    ])
    
    # Données
    for commande in commandes:
        writer.writerow([
            commande.numero_commande,
            commande.client.nom,
            commande.date_commande.strftime('%d/%m/%Y'),
            commande.date_livraison.strftime('%d/%m/%Y') if commande.date_livraison else '',
            commande.get_statut_display(),
            commande.entrepot.nom if commande.entrepot else '',
            f"{commande.sous_total:.2f}",
            f"{commande.montant_tva:.2f}",
            f"{commande.total:.2f}",
            commande.cree_par.username if commande.cree_par else ''
        ])
    
    return response


def exporter_commandes_pdf(commandes):
    """Exporter les commandes en format PDF"""
    from datetime import datetime
    
    buffer = BytesIO()
    
    # Créer le document PDF en paysage
    doc = SimpleDocTemplate(
        buffer,
        pagesize=unescape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title_style = styles['Heading1']
    title_style.alignment = 1  # Centré
    title = Paragraph(f"Liste des Commandes de Vente - {datetime.now().strftime('%d/%m/%Y')}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # Préparer les données du tableau
    data = [['N° Commande', 'Client', 'Date Commande', 'Date Livraison', 'Statut', 'Total']]
    
    for commande in commandes:
        data.append([
            commande.numero_commande,
            commande.client.nom[:25],  # Tronquer si trop long
            commande.date_commande.strftime('%d/%m/%Y'),
            commande.date_livraison.strftime('%d/%m/%Y') if commande.date_livraison else '',
            commande.get_statut_display(),
            f"{commande.total:,.0f} FCFA"
        ])
    
    # Ajouter ligne de total
    total_general = sum(c.total for c in commandes)
    data.append(['', '', '', '', 'TOTAL:', f"{total_general:,.0f} FCFA"])
    
    # Créer le tableau
    table = Table(data, colWidths=[4*cm, 5*cm, 3*cm, 3*cm, 3*cm, 3.5*cm])
    
    # Style du tableau
    table.setStyle(TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Corps du tableau
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        
        # Ligne de total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E5E7EB')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('ALIGN', (5, -1), (5, -1), 'RIGHT'),
    ]))
    
    elements.append(table)
    
    # Construire le PDF
    doc.build(elements)
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f"commandes_vente_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response# ========== FONCTION D'EXPORTATION POUR COMMANDES DE VENTE ==========

@login_required
def exporter_commandes_vente(request):
    """
    Vue pour exporter les commandes de vente en Excel, CSV ou PDF
    """
    format_export = request.GET.get('format', 'excel')
    filtre_statut = request.GET.get('statut', '')
    recherche = request.GET.get('recherche', '')
    ids = request.GET.get('ids', '')
    
    # Filtrer les commandes
    commandes = CommandeVente.objects.select_related('client', 'entrepot', 'cree_par')
    
    # Si des IDs spécifiques sont fournis (sélection)
    if ids:
        liste_ids = [int(id) for id in ids.split(',') if id.isdigit()]
        commandes = commandes.filter(id__in=liste_ids)
    else:
        # Sinon appliquer les filtres normaux
        if filtre_statut:
            commandes = commandes.filter(statut=filtre_statut)
        
        if recherche:
            commandes = commandes.filter(
                Q(numero_commande__icontains=recherche) |
                Q(client__nom__icontains=recherche)
            )
    
    commandes = commandes.order_by('-date_creation')
    
    # Appeler la fonction appropriée selon le format
    if format_export == 'excel':
        return exporter_commandes_excel(commandes)
    elif format_export == 'csv':
        return exporter_commandes_csv(commandes)
    elif format_export == 'pdf':
        return exporter_commandes_pdf(commandes)
    else:
        return HttpResponse("Format non supporté", status=400)


def exporter_commandes_excel(commandes):
    """Exporter les commandes en format Excel"""
    from datetime import datetime
    
    # Créer un workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes de Vente"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = [
        'N° Commande',
        'Client',
        'Date Commande',
        'Date Livraison',
        'Statut',
        'Entrepôt',
        'Sous-total',
        'TVA',
        'Total',
        'Créé par'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Données
    for row_idx, commande in enumerate(commandes, start=2):
        data_row = [
            commande.numero_commande,
            commande.client.nom,
            commande.date_commande.strftime('%d/%m/%Y'),
            commande.date_livraison.strftime('%d/%m/%Y') if commande.date_livraison else '',
            commande.get_statut_display(),
            commande.entrepot.nom if commande.entrepot else '',
            float(commande.sous_total),
            float(commande.montant_tva),
            float(commande.total),
            commande.cree_par.username if commande.cree_par else ''
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            # Alignement pour les montants
            if col_idx in [7, 8, 9]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 30, 15, 15, 15, 20, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Ligne de total
    total_row = len(commandes) + 2
    ws.cell(row=total_row, column=6).value = "TOTAL:"
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    ws.cell(row=total_row, column=7).value = sum(float(c.sous_total) for c in commandes)
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = '#,##0.00'
    ws.cell(row=total_row, column=8).value = sum(float(c.montant_tva) for c in commandes)
    ws.cell(row=total_row, column=8).font = Font(bold=True)
    ws.cell(row=total_row, column=8).number_format = '#,##0.00'
    ws.cell(row=total_row, column=9).value = sum(float(c.total) for c in commandes)
    ws.cell(row=total_row, column=9).font = Font(bold=True)
    ws.cell(row=total_row, column=9).number_format = '#,##0.00'
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Préparer la réponse
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"commandes_vente_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def exporter_commandes_csv(commandes):
    """Exporter les commandes en format CSV"""
    from datetime import datetime
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"commandes_vente_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Ajouter BOM UTF-8 pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # En-têtes
    writer.writerow([
        'N° Commande',
        'Client',
        'Date Commande',
        'Date Livraison',
        'Statut',
        'Entrepôt',
        'Sous-total',
        'TVA',
        'Total',
        'Créé par'
    ])
    
    # Données
    for commande in commandes:
        writer.writerow([
            commande.numero_commande,
            commande.client.nom,
            commande.date_commande.strftime('%d/%m/%Y'),
            commande.date_livraison.strftime('%d/%m/%Y') if commande.date_livraison else '',
            commande.get_statut_display(),
            commande.entrepot.nom if commande.entrepot else '',
            f"{commande.sous_total:.2f}",
            f"{commande.montant_tva:.2f}",
            f"{commande.total:.2f}",
            commande.cree_par.username if commande.cree_par else ''
        ])
    
    return response


def exporter_commandes_pdf(commandes):
    """Exporter les commandes en format PDF"""
    from datetime import datetime
    
    buffer = BytesIO()
    
    # Créer le document PDF en paysage
    doc = SimpleDocTemplate(
        buffer,
        pagesize=unescape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title_style = styles['Heading1']
    title_style.alignment = 1  # Centré
    title = Paragraph(f"Liste des Commandes de Vente - {datetime.now().strftime('%d/%m/%Y')}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # Préparer les données du tableau
    data = [['N° Commande', 'Client', 'Date Commande', 'Date Livraison', 'Statut', 'Total']]
    
    for commande in commandes:
        data.append([
            commande.numero_commande,
            commande.client.nom[:25],  # Tronquer si trop long
            commande.date_commande.strftime('%d/%m/%Y'),
            commande.date_livraison.strftime('%d/%m/%Y') if commande.date_livraison else '',
            commande.get_statut_display(),
            f"{commande.total:,.0f} FCFA"
        ])
    
    # Ajouter ligne de total
    total_general = sum(c.total for c in commandes)
    data.append(['', '', '', '', 'TOTAL:', f"{total_general:,.0f} FCFA"])
    
    # Créer le tableau
    table = Table(data, colWidths=[4*cm, 5*cm, 3*cm, 3*cm, 3*cm, 3.5*cm])
    
    # Style du tableau
    table.setStyle(TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Corps du tableau
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        
        # Ligne de total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E5E7EB')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('ALIGN', (5, -1), (5, -1), 'RIGHT'),
    ]))
    
    elements.append(table)
    
    # Construire le PDF
    doc.build(elements)
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f"commandes_vente_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response