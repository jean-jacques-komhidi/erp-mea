# comptabilite/views.py - VUES COMPLÈTES DU MODULE COMPTABILITÉ

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal

from .models import (
    PlanComptable, Exercice, Journal, Piece, Ecriture,
    Banque, MouvementBancaire, Budget
)
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ========== TABLEAU DE BORD COMPTABILITÉ ==========

@login_required
def tableau_bord_comptabilite(request):
    """
    Vue pour afficher le tableau de bord de la comptabilité
    """
    # Exercice en cours
    exercice_actuel = Exercice.objects.filter(
        est_cloture=False,
        date_debut__lte=timezone.now().date(),
        date_fin__gte=timezone.now().date()
    ).first()
    
    # Statistiques globales
    total_comptes = PlanComptable.objects.filter(est_actif=True).count()
    total_journaux = Journal.objects.filter(est_actif=True).count()
    total_pieces = Piece.objects.filter(est_validee=True).count()
    
    # Dernières pièces comptables
    dernieres_pieces = Piece.objects.select_related(
        'journal', 'exercice', 'cree_par'
    ).order_by('-date_creation')[:10]
    
    context = {
        'exercice_actuel': exercice_actuel,
        'total_comptes': total_comptes,
        'total_journaux': total_journaux,
        'total_pieces': total_pieces,
        'dernieres_pieces': dernieres_pieces,
    }
    
    return render(request, 'comptabilite/tableau_bord.jinja', context)


# ========== PLAN COMPTABLE ==========

@login_required
def liste_plan_comptable(request):
    """
    Vue pour afficher la liste des comptes du plan comptable SYSCOHADA
    """
    # Récupérer les paramètres de filtrage
    recherche = request.GET.get('recherche', '')
    type_compte = request.GET.get('type_compte', '')
    classe = request.GET.get('classe', '')
    statut = request.GET.get('statut', 'actif')
    
    # Filtrer les comptes
    comptes = PlanComptable.objects.all()
    
    if statut == 'actif':
        comptes = comptes.filter(est_actif=True)
    elif statut == 'inactif':
        comptes = comptes.filter(est_actif=False)
    
    if recherche:
        comptes = comptes.filter(
            Q(numero_compte__icontains=recherche) |
            Q(libelle__icontains=recherche)
        )
    
    if type_compte:
        comptes = comptes.filter(type_compte=type_compte)
    
    if classe:
        comptes = comptes.filter(numero_compte__startswith=classe)
    
    comptes = comptes.order_by('numero_compte')
    
    # Statistiques
    total_comptes = comptes.count()
    comptes_actifs = PlanComptable.objects.filter(est_actif=True).count()
    comptes_inactifs = PlanComptable.objects.filter(est_actif=False).count()
    
    # Répartition par type
    stats_par_type = []
    for type_code, type_label in PlanComptable.TYPES_COMPTE:
        nombre = PlanComptable.objects.filter(type_compte=type_code, est_actif=True).count()
        stats_par_type.append({
            'code': type_code,
            'label': type_label,
            'nombre': nombre
        })
    
    # Répartition par classe SYSCOHADA
    classes_syscohada = [
        ('1', 'Classe 1 - Comptes de ressources durables'),
        ('2', 'Classe 2 - Comptes d\'actif immobilisé'),
        ('3', 'Classe 3 - Comptes de stocks'),
        ('4', 'Classe 4 - Comptes de tiers'),
        ('5', 'Classe 5 - Comptes de trésorerie'),
        ('6', 'Classe 6 - Comptes de charges'),
        ('7', 'Classe 7 - Comptes de produits'),
        ('8', 'Classe 8 - Comptes des autres charges et produits'),
    ]
    
    stats_par_classe = []
    for classe_num, classe_label in classes_syscohada:
        nombre = PlanComptable.objects.filter(
            numero_compte__startswith=classe_num,
            est_actif=True
        ).count()
        stats_par_classe.append({
            'numero': classe_num,
            'label': classe_label,
            'nombre': nombre
        })
    
    context = {
        'comptes': comptes,
        'recherche': recherche,
        'type_compte': type_compte,
        'classe': classe,
        'statut': statut,
        'total_comptes': total_comptes,
        'comptes_actifs': comptes_actifs,
        'comptes_inactifs': comptes_inactifs,
        'stats_par_type': stats_par_type,
        'stats_par_classe': stats_par_classe,
        'types_compte': PlanComptable.TYPES_COMPTE,
        'classes_syscohada': classes_syscohada,
    }
    
    return render(request, 'comptabilite/liste_plan_comptable.jinja', context)


@login_required
def creer_compte(request):
    """Vue pour créer un nouveau compte comptable"""
    if request.method == 'POST':
        numero_compte = request.POST.get('numero_compte', '').strip()
        libelle = request.POST.get('libelle', '').strip()
        type_compte = request.POST.get('type_compte')
        compte_parent_id = request.POST.get('compte_parent')
        
        if not numero_compte or not libelle or not type_compte:
            messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
            return redirect('comptabilite:creer_compte')
        
        if PlanComptable.objects.filter(numero_compte=numero_compte).exists():
            messages.error(request, f'Le compte {numero_compte} existe déjà.')
            return redirect('comptabilite:creer_compte')
        
        compte = PlanComptable.objects.create(
            numero_compte=numero_compte,
            libelle=libelle,
            type_compte=type_compte,
            compte_parent_id=compte_parent_id if compte_parent_id else None,
            est_actif=True
        )
        
        messages.success(request, f'Compte {compte.numero_compte} - {compte.libelle} créé avec succès !')
        return redirect('comptabilite:details_compte', pk=compte.pk)
    
    comptes_parents = PlanComptable.objects.filter(est_actif=True).order_by('numero_compte')
    
    context = {
        'comptes_parents': comptes_parents,
        'types_compte': PlanComptable.TYPES_COMPTE,
    }
    
    return render(request, 'comptabilite/formulaire_compte.jinja', context)


@login_required
def details_compte(request, pk):
    """Vue pour afficher les détails d'un compte comptable"""
    compte = get_object_or_404(PlanComptable, pk=pk)
    
    ecritures = Ecriture.objects.filter(compte=compte).select_related(
        'piece', 'piece__journal'
    ).order_by('-piece__date_piece')[:50]
    
    debits = Ecriture.objects.filter(compte=compte, piece__est_validee=True).aggregate(
        total=Sum('debit')
    )['total'] or Decimal('0')
    
    credits = Ecriture.objects.filter(compte=compte, piece__est_validee=True).aggregate(
        total=Sum('credit')
    )['total'] or Decimal('0')
    
    if compte.type_compte in ['ACTIF', 'CHARGE']:
        solde = debits - credits
        sens_solde = 'Débiteur' if solde >= 0 else 'Créditeur'
    else:
        solde = credits - debits
        sens_solde = 'Créditeur' if solde >= 0 else 'Débiteur'
    
    nombre_ecritures = ecritures.count()
    
    context = {
        'compte': compte,
        'ecritures': ecritures,
        'solde': abs(solde),
        'sens_solde': sens_solde,
        'debits': debits,
        'credits': credits,
        'nombre_ecritures': nombre_ecritures,
    }
    
    return render(request, 'comptabilite/details_compte.jinja', context)


@login_required
def exporter_plan_comptable(request):
    """Vue pour exporter le plan comptable en Excel"""
    recherche = request.GET.get('recherche', '')
    type_compte = request.GET.get('type_compte', '')
    classe = request.GET.get('classe', '')
    statut = request.GET.get('statut', 'actif')
    
    comptes = PlanComptable.objects.all()
    
    if statut == 'actif':
        comptes = comptes.filter(est_actif=True)
    elif statut == 'inactif':
        comptes = comptes.filter(est_actif=False)
    
    if recherche:
        comptes = comptes.filter(
            Q(numero_compte__icontains=recherche) |
            Q(libelle__icontains=recherche)
        )
    
    if type_compte:
        comptes = comptes.filter(type_compte=type_compte)
    
    if classe:
        comptes = comptes.filter(numero_compte__startswith=classe)
    
    comptes = comptes.order_by('numero_compte')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan Comptable"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['Numéro de compte', 'Libellé', 'Type', 'Compte parent', 'Statut']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row_idx, compte in enumerate(comptes, start=2):
        data_row = [
            compte.numero_compte,
            compte.libelle,
            compte.get_type_compte_display(),
            f"{compte.compte_parent.numero_compte} - {compte.compte_parent.libelle}" if compte.compte_parent else '',
            'Actif' if compte.est_actif else 'Inactif'
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
    
    column_widths = [20, 40, 15, 35, 12]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"plan_comptable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# ========== EXERCICES ==========

@login_required
def liste_exercices(request):
    """Vue pour lister les exercices comptables"""
    exercices = Exercice.objects.all().order_by('-date_debut')
    
    context = {
        'exercices': exercices,
    }
    
    return render(request, 'comptabilite/liste_exercices.jinja', context)


@login_required
def creer_exercice(request):
    """Vue pour créer un nouvel exercice"""
    messages.info(request, 'Fonctionnalité en cours de développement.')
    return redirect('comptabilite:liste_exercices')


# ========== JOURNAUX ==========

@login_required
def liste_journaux(request):
    """Vue pour lister les journaux comptables"""
    journaux = Journal.objects.all()
    
    context = {
        'journaux': journaux,
    }
    
    return render(request, 'comptabilite/liste_journaux.jinja', context)


@login_required
def creer_journal(request):
    """Vue pour créer un nouveau journal"""
    messages.info(request, 'Fonctionnalité en cours de développement.')
    return redirect('comptabilite:liste_journaux')


# ========== PIÈCES COMPTABLES ==========

@login_required
def liste_pieces(request):
    """Vue pour lister les pièces comptables"""
    pieces = Piece.objects.select_related(
        'journal', 'exercice', 'cree_par'
    ).order_by('-date_piece')
    
    context = {
        'pieces': pieces,
    }
    
    return render(request, 'comptabilite/liste_pieces.jinja', context)


@login_required
def creer_piece(request):
    """Vue pour créer une nouvelle pièce comptable"""
    messages.info(request, 'Fonctionnalité en cours de développement.')
    return redirect('comptabilite:liste_pieces')


@login_required
def details_piece(request, pk):
    """Vue pour afficher les détails d'une pièce"""
    piece = get_object_or_404(Piece, pk=pk)
    
    context = {
        'piece': piece,
    }
    
    return render(request, 'comptabilite/details_piece.jinja', context)


@login_required
def valider_piece(request, pk):
    """Vue pour valider une pièce comptable"""
    messages.info(request, 'Fonctionnalité en cours de développement.')
    return redirect('comptabilite:details_piece', pk=pk)


# ========== BANQUES ==========

@login_required
def liste_banques(request):
    """Vue pour lister les comptes bancaires"""
    banques = Banque.objects.filter(est_actif=True)
    
    context = {
        'banques': banques,
    }
    
    return render(request, 'comptabilite/liste_banques.jinja', context)


@login_required
def creer_banque(request):
    """Vue pour créer un nouveau compte bancaire"""
    messages.info(request, 'Fonctionnalité en cours de développement.')
    return redirect('comptabilite:liste_banques')


@login_required
def details_banque(request, pk):
    """Vue pour afficher les détails d'un compte bancaire"""
    banque = get_object_or_404(Banque, pk=pk)
    
    context = {
        'banque': banque,
    }
    
    return render(request, 'comptabilite/details_banque.jinja', context)


@login_required
def creer_mouvement_bancaire(request):
    """Vue pour créer un mouvement bancaire"""
    messages.info(request, 'Fonctionnalité en cours de développement.')
    return redirect('comptabilite:liste_banques')


# ========== BUDGETS ==========

@login_required
def liste_budgets(request):
    """Vue pour lister les budgets"""
    budgets = Budget.objects.select_related('exercice', 'compte').all()
    
    context = {
        'budgets': budgets,
    }
    
    return render(request, 'comptabilite/liste_budgets.jinja', context)


@login_required
def creer_budget(request):
    """Vue pour créer un budget"""
    messages.info(request, 'Fonctionnalité en cours de développement.')
    return redirect('comptabilite:liste_budgets')


# ========== RAPPORTS ==========

@login_required
def bilan(request):
    """Vue pour afficher le bilan"""
    messages.info(request, 'Fonctionnalité en cours de développement.')
    return redirect('comptabilite:tableau_bord')


@login_required
def compte_resultat(request):
    """Vue pour afficher le compte de résultat"""
    messages.info(request, 'Fonctionnalité en cours de développement.')
    return redirect('comptabilite:tableau_bord')