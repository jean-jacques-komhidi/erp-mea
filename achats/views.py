# achats/views.py - Vues complètes du module achats

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Sum, Count, F, Avg
from django.http import JsonResponse
from django.utils import timezone
from datetime import datetime, timedelta, date
from decimal import Decimal

from .models import CommandeAchat, LigneCommandeAchat, PaiementFournisseur
from stock.models import MouvementStock, Produit, Entrepot
from base.models import Fournisseur
from .forms import (
    CommandeAchatForm, LigneCommandeAchatFormSet, 
    RecevoirCommandeForm, AnnulerCommandeForm
)


# ========== LISTE DES COMMANDES D'ACHAT ==========

@login_required
def liste_commandes_achat(request):
    """Vue pour afficher la liste des commandes d'achat"""
    filtre_statut = request.GET.get('statut', '')
    recherche = request.GET.get('recherche', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    commandes = CommandeAchat.objects.select_related('fournisseur', 'entrepot', 'cree_par')
    
    # Filtres
    if filtre_statut:
        commandes = commandes.filter(statut=filtre_statut)
    
    if recherche:
        commandes = commandes.filter(
            Q(numero_commande__icontains=recherche) |
            Q(fournisseur__nom__icontains=recherche) |
            Q(fournisseur__code__icontains=recherche)
        )
    
    if date_debut:
        commandes = commandes.filter(date_commande__gte=date_debut)
    
    if date_fin:
        commandes = commandes.filter(date_commande__lte=date_fin)
    
    commandes = commandes.order_by('-date_creation')
    
    # Statistiques
    aujourdhui = timezone.now().date()
    total_montant = commandes.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
    
    commandes_en_attente = commandes.filter(statut='BROUILLON').count()
    commandes_confirmees = commandes.filter(statut='CONFIRMEE').count()
    commandes_en_cours = commandes.filter(statut__in=['CONFIRMEE', 'ENVOYEE']).count()
    
    # Commandes en retard de livraison
    commandes_en_retard = commandes.filter(
        date_livraison_prevue__lt=aujourdhui,
        statut__in=['CONFIRMEE', 'ENVOYEE']
    ).count()
    
    # Commandes du mois en cours
    commandes_ce_mois = commandes.filter(
        date_creation__year=aujourdhui.year,
        date_creation__month=aujourdhui.month
    ).count()
    
    contexte = {
        'commandes': commandes,
        'filtre_statut': filtre_statut,
        'recherche': recherche,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'statuts': CommandeAchat.STATUTS,
        'aujourdhui': aujourdhui,
        'total_montant': total_montant,
        'commandes_en_attente': commandes_en_attente,
        'commandes_confirmees': commandes_confirmees,
        'commandes_en_cours': commandes_en_cours,
        'commandes_en_retard': commandes_en_retard,
        'commandes_ce_mois': commandes_ce_mois,
    }
    
    return render(request, 'achats/liste_commandes.jinja', contexte)


# ========== CRÉER UNE COMMANDE D'ACHAT ==========

@login_required
@transaction.atomic
def creer_commande_achat(request):
    """Vue pour créer une nouvelle commande d'achat"""
    
    if request.method == 'POST':
        # Récupérer les données du formulaire
        fournisseur_id = request.POST.get('fournisseur')
        entrepot_id = request.POST.get('entrepot')
        date_livraison_prevue = request.POST.get('date_livraison_prevue')
        notes = request.POST.get('notes', '')
        
        # Valider les données
        if not fournisseur_id or not entrepot_id:
            messages.error(request, 'Veuillez sélectionner un fournisseur et un entrepôt.')
            return redirect('achats:creer_commande_achat')
        
        # Créer la commande
        commande = CommandeAchat.objects.create(
            fournisseur_id=fournisseur_id,
            entrepot_id=entrepot_id,
            date_livraison_prevue=date_livraison_prevue if date_livraison_prevue else None,
            notes=notes,
            cree_par=request.user,
            statut='BROUILLON'
        )
        
        # Créer les lignes de commande
        produits_ids = request.POST.getlist('produit[]')
        quantites = request.POST.getlist('quantite[]')
        prix_unitaires = request.POST.getlist('prix_unitaire[]')
        taux_tvas = request.POST.getlist('taux_tva[]')
        
        lignes_creees = 0
        for i, produit_id in enumerate(produits_ids):
            if produit_id and quantites[i]:
                try:
                    produit = Produit.objects.get(pk=produit_id)
                    quantite = int(quantites[i])
                    prix_unitaire = Decimal(prix_unitaires[i]) if prix_unitaires[i] else produit.prix_achat
                    taux_tva = Decimal(taux_tvas[i]) if taux_tvas[i] else produit.taux_tva
                    
                    LigneCommandeAchat.objects.create(
                        commande=commande,
                        produit=produit,
                        quantite=quantite,
                        prix_unitaire=prix_unitaire,
                        taux_tva=taux_tva
                    )
                    lignes_creees += 1
                except (Produit.DoesNotExist, ValueError, Decimal.InvalidOperation) as e:
                    messages.warning(request, f'Ligne {i+1} ignorée : {str(e)}')
                    continue
        
        if lignes_creees == 0:
            commande.delete()
            messages.error(request, 'Veuillez ajouter au moins un produit à la commande.')
            return redirect('achats:creer_commande_achat')
        
        # Calculer les totaux
        commande.calculer_totaux()
        
        messages.success(request, f'Commande {commande.numero_commande} créée avec succès avec {lignes_creees} ligne(s)!')
        return redirect('achats:details_commande_achat', pk=commande.pk)
    
    # GET - Afficher le formulaire
    fournisseurs = Fournisseur.objects.filter(est_actif=True).order_by('nom')
    entrepots = Entrepot.objects.filter(est_actif=True).order_by('nom')
    produits = Produit.objects.filter(est_actif=True).order_by('nom')
    
    # Date de livraison par défaut : 7 jours
    date_livraison_defaut = (timezone.now().date() + timedelta(days=7)).strftime('%Y-%m-%d')
    
    contexte = {
        'fournisseurs': fournisseurs,
        'entrepots': entrepots,
        'produits': produits,
        'action': 'Créer',
        'date_livraison_defaut': date_livraison_defaut,
    }
    
    return render(request, 'achats/formulaire_commande.jinja', contexte)


# ========== DÉTAILS D'UNE COMMANDE ==========

@login_required
def details_commande_achat(request, pk):
    """Vue pour afficher les détails d'une commande d'achat"""
    commande = get_object_or_404(
        CommandeAchat.objects.select_related('fournisseur', 'entrepot', 'cree_par', 'piece_comptable'),
        pk=pk
    )
    
    lignes = commande.lignecommandeachat_set.select_related('produit').all()
    paiements = commande.paiements.select_related('utilisateur').order_by('-date_paiement')
    
    # Mouvements de stock liés à cette commande
    mouvements = MouvementStock.objects.filter(
        reference=commande.numero_commande
    ).select_related('produit', 'entrepot', 'utilisateur').order_by('-date')
    
    # Calculs
    total_paye = paiements.aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
    solde_restant = commande.total - total_paye
    
    contexte = {
        'commande': commande,
        'lignes': lignes,
        'paiements': paiements,
        'mouvements': mouvements,
        'total_paye': total_paye,
        'solde_restant': solde_restant,
        'aujourdhui': timezone.now().date(),
        'taux_reception': commande.taux_reception(),
    }
    
    return render(request, 'achats/details_commande.jinja', contexte)


# ========== MODIFIER UNE COMMANDE ==========

@login_required
@transaction.atomic
def modifier_commande_achat(request, pk):
    """Vue pour modifier une commande d'achat"""
    commande = get_object_or_404(CommandeAchat, pk=pk)
    
    # Vérifier si la commande peut être modifiée
    if not commande.peut_etre_modifiee():
        messages.error(request, f'Cette commande ne peut plus être modifiée (statut : {commande.get_statut_display()}).')
        return redirect('achats:details_commande_achat', pk=pk)
    
    if request.method == 'POST':
        # Mettre à jour la commande
        commande.fournisseur_id = request.POST.get('fournisseur')
        commande.entrepot_id = request.POST.get('entrepot')
        date_livraison = request.POST.get('date_livraison_prevue')
        commande.date_livraison_prevue = date_livraison if date_livraison else None
        commande.notes = request.POST.get('notes', '')
        commande.save()
        
        # Supprimer les anciennes lignes et créer les nouvelles
        commande.lignecommandeachat_set.all().delete()
        
        produits_ids = request.POST.getlist('produit[]')
        quantites = request.POST.getlist('quantite[]')
        prix_unitaires = request.POST.getlist('prix_unitaire[]')
        taux_tvas = request.POST.getlist('taux_tva[]')
        
        lignes_creees = 0
        for i, produit_id in enumerate(produits_ids):
            if produit_id and quantites[i]:
                try:
                    produit = Produit.objects.get(pk=produit_id)
                    LigneCommandeAchat.objects.create(
                        commande=commande,
                        produit=produit,
                        quantite=int(quantites[i]),
                        prix_unitaire=Decimal(prix_unitaires[i]) if prix_unitaires[i] else produit.prix_achat,
                        taux_tva=Decimal(taux_tvas[i]) if taux_tvas[i] else produit.taux_tva
                    )
                    lignes_creees += 1
                except (Produit.DoesNotExist, ValueError, Decimal.InvalidOperation):
                    continue
        
        if lignes_creees == 0:
            messages.error(request, 'Veuillez ajouter au moins un produit à la commande.')
            return redirect('achats:modifier_commande_achat', pk=pk)
        
        commande.calculer_totaux()
        
        messages.success(request, f'Commande {commande.numero_commande} modifiée avec succès!')
        return redirect('achats:details_commande_achat', pk=pk)
    
    # GET - Afficher le formulaire
    fournisseurs = Fournisseur.objects.filter(est_actif=True).order_by('nom')
    entrepots = Entrepot.objects.filter(est_actif=True).order_by('nom')
    produits = Produit.objects.filter(est_actif=True).order_by('nom')
    lignes = commande.lignecommandeachat_set.select_related('produit').all()
    
    contexte = {
        'commande': commande,
        'lignes': lignes,
        'fournisseurs': fournisseurs,
        'entrepots': entrepots,
        'produits': produits,
        'action': 'Modifier',
    }
    
    return render(request, 'achats/formulaire_commande.jinja', contexte)


# ========== CONFIRMER UNE COMMANDE ==========

@login_required
def confirmer_commande_achat(request, pk):
    """Vue pour confirmer une commande d'achat"""
    commande = get_object_or_404(CommandeAchat, pk=pk)
    
    if not commande.peut_etre_confirmee():
        messages.error(request, 'Cette commande ne peut pas être confirmée.')
        return redirect('achats:details_commande_achat', pk=pk)
    
    if request.method == 'POST':
        commande.statut = 'CONFIRMEE'
        commande.save()
        messages.success(request, f'Commande {commande.numero_commande} confirmée avec succès!')
        return redirect('achats:details_commande_achat', pk=pk)
    
    contexte = {'commande': commande}
    return render(request, 'achats/confirmer_commande.jinja', contexte)


# ========== ENVOYER UNE COMMANDE AU FOURNISSEUR ==========

@login_required
def envoyer_commande_achat(request, pk):
    """Vue pour marquer une commande comme envoyée au fournisseur"""
    commande = get_object_or_404(CommandeAchat, pk=pk)
    
    if not commande.peut_etre_envoyee():
        messages.error(request, 'Cette commande ne peut pas être envoyée (elle doit être confirmée d\'abord).')
        return redirect('achats:details_commande_achat', pk=pk)
    
    if request.method == 'POST':
        commande.statut = 'ENVOYEE'
        commande.save()
        
        messages.success(request, f'Commande {commande.numero_commande} marquée comme envoyée au fournisseur {commande.fournisseur.nom}!')
        return redirect('achats:details_commande_achat', pk=pk)
    
    contexte = {'commande': commande}
    return render(request, 'achats/envoyer_commande.jinja', contexte)


# ========== RECEVOIR UNE COMMANDE ==========

@login_required
@transaction.atomic
def recevoir_commande_achat(request, pk):
    """Vue pour recevoir une commande d'achat"""
    commande = get_object_or_404(CommandeAchat, pk=pk)
    
    if not commande.peut_etre_recue():
        messages.error(request, 'Cette commande ne peut pas être reçue.')
        return redirect('achats:details_commande_achat', pk=pk)
    
    lignes = commande.lignecommandeachat_set.select_related('produit').all()
    
    if request.method == 'POST':
        # Récupérer les données du formulaire (compatible avec le template)
        ligne_ids = request.POST.getlist('ligne_id[]')
        quantites_recues = request.POST.getlist('quantite_recue[]')
        notes_reception = request.POST.get('notes_reception', '')
        
        # Vérifier qu'il y a au moins une quantité à réceptionner
        total_recu = sum([int(q) if q else 0 for q in quantites_recues])
        
        if total_recu == 0:
            messages.error(request, "Veuillez saisir au moins une quantité à réceptionner.")
            return redirect('achats:recevoir_commande_achat', pk=pk)
        
        # Traiter chaque ligne
        mouvements_crees = 0
        for i, ligne_id in enumerate(ligne_ids):
            quantite_recue = int(quantites_recues[i]) if quantites_recues[i] else 0
            
            if quantite_recue > 0:
                ligne = LigneCommandeAchat.objects.get(pk=ligne_id)
                
                # Vérifier que la quantité ne dépasse pas le restant
                quantite_restante = ligne.quantite - ligne.quantite_recue
                if quantite_recue > quantite_restante:
                    messages.error(
                        request, 
                        f"La quantité reçue pour {ligne.produit.nom} dépasse la quantité restante ({quantite_restante})."
                    )
                    return redirect('achats:recevoir_commande_achat', pk=pk)
                
                # Mettre à jour la quantité reçue
                ligne.quantite_recue += quantite_recue
                ligne.save()
                
                # Créer le mouvement de stock (ENTRÉE)
                MouvementStock.objects.create(
                    produit=ligne.produit,
                    entrepot=commande.entrepot,
                    type_mouvement='ENTREE',
                    quantite=quantite_recue,
                    reference=commande.numero_commande,
                    notes=f'Réception commande achat {commande.numero_commande} - {commande.fournisseur.nom}',
                    utilisateur=request.user
                )
                
                mouvements_crees += 1
        
        # Mettre à jour la date de réception si c'est la première fois
        if not commande.date_reception:
            commande.date_reception = timezone.now().date()
        
        # Vérifier si toutes les lignes sont complètement reçues
        if commande.est_completement_recue():
            commande.statut = 'RECUE'
            
            # Générer l'écriture comptable automatiquement
            try:
                from comptabilite.models import Journal, Exercice
                
                journal_achat = Journal.objects.filter(type_journal='ACHAT', est_actif=True).first()
                aujourdhui = timezone.now().date()
                exercice = Exercice.objects.filter(
                    est_cloture=False,
                    date_debut__lte=aujourdhui,
                    date_fin__gte=aujourdhui
                ).first()
                
                if journal_achat and exercice:
                    piece = commande.generer_ecriture_comptable(journal_achat, exercice)
                    messages.success(request, f'Écriture comptable {piece.numero_piece} générée automatiquement!')
                else:
                    messages.warning(request, 'Impossible de générer l\'écriture comptable : journal ou exercice introuvable.')
            except Exception as e:
                messages.warning(request, f'Erreur lors de la génération de l\'écriture comptable : {str(e)}')
        
        # Ajouter les notes de réception
        if notes_reception:
            if commande.notes:
                commande.notes += f"\n\n[Réception {timezone.now().strftime('%d/%m/%Y')}]: {notes_reception}"
            else:
                commande.notes = f"[Réception {timezone.now().strftime('%d/%m/%Y')}]: {notes_reception}"
        
        commande.save()
        
        messages.success(
            request, 
            f'Réception enregistrée ! {mouvements_crees} mouvement(s) de stock créé(s). Le stock a été mis à jour automatiquement.'
        )
        return redirect('achats:details_commande_achat', pk=pk)
    
    # GET request
    context = {
        'commande': commande,
        'lignes': lignes,
        'aujourdhui': timezone.now().date(),
    }
    
    return render(request, 'achats/recevoir_commande.jinja', context)


# ========== ANNULER UNE COMMANDE ==========

@login_required
@transaction.atomic
def annuler_commande_achat(request, pk):
    """Vue pour annuler une commande d'achat"""
    commande = get_object_or_404(CommandeAchat, pk=pk)
    
    if not commande.peut_etre_annulee():
        messages.error(request, 'Cette commande ne peut pas être annulée.')
        return redirect('achats:details_commande_achat', pk=pk)
    
    if request.method == 'POST':
        raison = request.POST.get('raison', '')
        
        if not raison:
            messages.error(request, 'Veuillez indiquer la raison de l\'annulation.')
            return redirect('achats:annuler_commande_achat', pk=pk)
        
        commande.statut = 'ANNULEE'
        commande.raison_annulation = raison
        commande.save()
        
        messages.success(request, f'Commande {commande.numero_commande} annulée.')
        return redirect('achats:details_commande_achat', pk=pk)
    
    contexte = {'commande': commande}
    return render(request, 'achats/annuler_commande.jinja', contexte)


# ========== SUPPRIMER UNE COMMANDE ==========

@login_required
@transaction.atomic
def supprimer_commande_achat(request, pk):
    """Vue pour supprimer une commande d'achat"""
    commande = get_object_or_404(CommandeAchat, pk=pk)
    
    # Seules les commandes BROUILLON peuvent être supprimées
    if commande.statut != 'BROUILLON':
        messages.error(request, 'Seules les commandes au statut "Brouillon" peuvent être supprimées.')
        return redirect('achats:details_commande_achat', pk=pk)
    
    if request.method == 'POST':
        numero = commande.numero_commande
        commande.delete()
        messages.success(request, f'Commande {numero} supprimée avec succès.')
        return redirect('achats:liste_commandes_achat')
    
    contexte = {'commande': commande}
    return render(request, 'achats/confirmer_suppression_commande.jinja', contexte)


# ========== HISTORIQUE DES ACHATS ==========

@login_required
def historique_achats(request):
    """Vue pour afficher l'historique et les statistiques des achats"""
    
    # Récupérer les paramètres de filtrage
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    fournisseur_id = request.GET.get('fournisseur')
    
    # Date par défaut (30 derniers jours)
    if not date_debut:
        date_debut = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_fin:
        date_fin = timezone.now().strftime('%Y-%m-%d')
    
    # Filtrer les commandes
    commandes = CommandeAchat.objects.all()
    
    if date_debut:
        commandes = commandes.filter(date_commande__gte=date_debut)
    if date_fin:
        commandes = commandes.filter(date_commande__lte=date_fin)
    if fournisseur_id:
        commandes = commandes.filter(fournisseur_id=fournisseur_id)
    
    # Statistiques globales
    stats_globales = {
        'nombre_commandes': commandes.count(),
        'montant_total': commandes.aggregate(total=Sum('total'))['total'] or 0,
        'nombre_fournisseurs': commandes.values('fournisseur').distinct().count(),
        'montant_moyen': commandes.aggregate(moyenne= Avg('total'))['moyenne'] or 0,
        'taux_reception': 0,
        'commandes_retard': 0,
        'taux_confirmation': 0,
    }
    
    # Calculer les taux
    if stats_globales['nombre_commandes'] > 0:
        commandes_recues = commandes.filter(statut__in=['RECUE', 'FACTUREE']).count()
        stats_globales['taux_reception'] = (commandes_recues / stats_globales['nombre_commandes']) * 100
        
        aujourd_hui = timezone.now().date()
        commandes_retard = commandes.filter(
            date_livraison_prevue__lt=aujourd_hui
        ).exclude(statut__in=['RECUE', 'FACTUREE', 'ANNULEE']).count()
        stats_globales['commandes_retard'] = commandes_retard
        
        commandes_confirmees = commandes.exclude(statut='BROUILLON').count()
        stats_globales['taux_confirmation'] = (commandes_confirmees / stats_globales['nombre_commandes']) * 100
    
    # Répartition par statut
    stats_par_statut = []
    statuts = ['BROUILLON', 'CONFIRMEE', 'ENVOYEE', 'RECUE', 'FACTUREE', 'ANNULEE']
    
    for statut in statuts:
        commandes_statut = commandes.filter(statut=statut)
        nombre = commandes_statut.count()
        montant = commandes_statut.aggregate(total=Sum('total'))['total'] or 0
        pourcentage = (nombre / stats_globales['nombre_commandes'] * 100) if stats_globales['nombre_commandes'] > 0 else 0
        
        if nombre > 0:  # N'ajouter que les statuts qui ont des commandes
            stats_par_statut.append({
                'statut': statut,
                'nombre': nombre,
                'montant': montant,
                'pourcentage': pourcentage
            })
    
    # Top 10 fournisseurs
    top_fournisseurs = commandes.values(
        'fournisseur__nom'
    ).annotate(
        nombre=Count('id'),
        montant=Sum('total')
    ).order_by('-montant')[:10]
    
    # Top 10 produits
    top_produits = LigneCommandeAchat.objects.filter(
        commande__in=commandes
    ).values(
        'produit__nom',
        'produit__code'
    ).annotate(
        quantite_totale=Sum('quantite'),
        montant_total=Sum(F('quantite') * F('prix_unitaire'))
    ).order_by('-montant_total')[:10]
    
    # Évolution mensuelle (6 derniers mois)
    evolution_mensuelle = []
    for i in range(6):
        mois_debut = (timezone.now() - timedelta(days=30*i)).replace(day=1)
        if i == 0:
            mois_fin = timezone.now()
        else:
            mois_fin = mois_debut.replace(day=1) + timedelta(days=32)
            mois_fin = mois_fin.replace(day=1) - timedelta(days=1)
        
        commandes_mois = commandes.filter(
            date_commande__gte=mois_debut,
            date_commande__lte=mois_fin
        )
        
        nombre = commandes_mois.count()
        montant = commandes_mois.aggregate(total=Sum('total'))['total'] or 0
        montant_moyen = montant / nombre if nombre > 0 else 0
        
        # Taux de réception du mois
        recues = commandes_mois.filter(statut__in=['RECUE', 'FACTUREE']).count()
        taux_reception = (recues / nombre * 100) if nombre > 0 else 0
        
        # Évolution par rapport au mois précédent
        evolution = 0
        if i < 5 and len(evolution_mensuelle) > 0:
            mois_precedent = evolution_mensuelle[-1]['montant']
            if mois_precedent > 0:
                evolution = ((montant - mois_precedent) / mois_precedent) * 100
        
        evolution_mensuelle.insert(0, {
            'mois': mois_debut.strftime('%B %Y'),
            'nombre': nombre,
            'montant': montant,
            'montant_moyen': montant_moyen,
            'taux_reception': taux_reception,
            'evolution': evolution
        })
    
    # Liste des fournisseurs pour le filtre
    fournisseurs = Fournisseur.objects.all().order_by('nom')
    
    context = {
        'stats_globales': stats_globales,
        'stats_par_statut': stats_par_statut,
        'top_fournisseurs': top_fournisseurs,
        'top_produits': top_produits,
        'evolution_mensuelle': evolution_mensuelle,
        'fournisseurs': fournisseurs,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    
    return render(request, 'achats/historique_achats.jinja', context)

# ========== API AJAX POUR OBTENIR LE PRIX D'ACHAT D'UN PRODUIT ==========

@login_required
def obtenir_prix_produit(request, pk):
    """API AJAX pour obtenir le prix d'achat d'un produit"""
    try:
        produit = Produit.objects.get(pk=pk)
        return JsonResponse({
            'succes': True,
            'prix_achat': float(produit.prix_achat),
            'taux_tva': float(produit.taux_tva),
            'stock_actuel': produit.stock_actuel,
            'code': produit.code,
            'unite': produit.unite,
        })
    except Produit.DoesNotExist:
        return JsonResponse({
            'succes': False,
            'erreur': 'Produit non trouvé'
        }, status=404)
    
# ========== FONCTION D'EXPORTATION POUR COMMANDES D'ACHAT ==========
# À AJOUTER à la fin de achats/views.py

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
import csv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import cm

@login_required
def exporter_commandes_achat(request):
    """
    Vue pour exporter les commandes d'achat en Excel, CSV ou PDF
    """
    format_export = request.GET.get('format', 'excel')
    filtre_statut = request.GET.get('statut', '')
    recherche = request.GET.get('recherche', '')
    ids = request.GET.get('ids', '')
    
    # Filtrer les commandes
    commandes = CommandeAchat.objects.select_related('fournisseur', 'entrepot', 'cree_par')
    
    # Si des IDs spécifiques sont fournis (sélection)
    if ids:
        liste_ids = [int(id) for id in ids.split(',') if id.isdigit()]
        commandes = commandes.filter(id__in=liste_ids)
    else:
        # Sinon appliquer les filtres normaux
        if filtre_statut:
            commandes = commandes.filter(statut=filtre_statut)
        
        if recherche:
            commandes = commandes.filter(
                Q(numero_commande__icontains=recherche) |
                Q(fournisseur__nom__icontains=recherche) |
                Q(reference__icontains=recherche)
            )
    
    commandes = commandes.order_by('-date_creation')
    
    # Appeler la fonction appropriée selon le format
    if format_export == 'excel':
        return exporter_commandes_achat_excel(commandes)
    elif format_export == 'csv':
        return exporter_commandes_achat_csv(commandes)
    elif format_export == 'pdf':
        return exporter_commandes_achat_pdf(commandes)
    else:
        return HttpResponse("Format non supporté", status=400)


def exporter_commandes_achat_excel(commandes):
    """Exporter les commandes d'achat en format Excel"""
    from datetime import datetime
    
    # Créer un workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes d'Achat"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = [
        'N° Commande',
        'Référence',
        'Fournisseur',
        'Date Commande',
        'Date Livraison',
        'Statut',
        'Entrepôt',
        'Sous-total',
        'TVA',
        'Total',
        'Créé par'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Données
    for row_idx, commande in enumerate(commandes, start=2):
        data_row = [
            commande.numero_commande,
            commande.reference or '',
            commande.fournisseur.nom,
            commande.date_commande.strftime('%d/%m/%Y'),
            commande.date_livraison_prevue.strftime('%d/%m/%Y') if commande.date_livraison_prevue else '',
            commande.get_statut_display(),
            commande.entrepot.nom if commande.entrepot else '',
            float(commande.sous_total),
            float(commande.montant_tva),
            float(commande.total),
            commande.cree_par.username if commande.cree_par else ''
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            # Alignement pour les montants
            if col_idx in [8, 9, 10]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 15, 30, 15, 15, 15, 20, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Ligne de total
    total_row = len(commandes) + 2
    ws.cell(row=total_row, column=7).value = "TOTAL:"
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=8).value = sum(float(c.sous_total) for c in commandes)
    ws.cell(row=total_row, column=8).font = Font(bold=True)
    ws.cell(row=total_row, column=8).number_format = '#,##0.00'
    ws.cell(row=total_row, column=9).value = sum(float(c.montant_tva) for c in commandes)
    ws.cell(row=total_row, column=9).font = Font(bold=True)
    ws.cell(row=total_row, column=9).number_format = '#,##0.00'
    ws.cell(row=total_row, column=10).value = sum(float(c.total) for c in commandes)
    ws.cell(row=total_row, column=10).font = Font(bold=True)
    ws.cell(row=total_row, column=10).number_format = '#,##0.00'
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Préparer la réponse
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"commandes_achat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def exporter_commandes_achat_csv(commandes):
    """Exporter les commandes d'achat en format CSV"""
    from datetime import datetime
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"commandes_achat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Ajouter BOM UTF-8 pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # En-têtes
    writer.writerow([
        'N° Commande',
        'Référence',
        'Fournisseur',
        'Date Commande',
        'Date Livraison',
        'Statut',
        'Entrepôt',
        'Sous-total',
        'TVA',
        'Total',
        'Créé par'
    ])
    
    # Données
    for commande in commandes:
        writer.writerow([
            commande.numero_commande,
            commande.reference or '',
            commande.fournisseur.nom,
            commande.date_commande.strftime('%d/%m/%Y'),
            commande.date_livraison_prevue.strftime('%d/%m/%Y') if commande.date_livraison_prevue else '',
            commande.get_statut_display(),
            commande.entrepot.nom if commande.entrepot else '',
            f"{commande.sous_total:.2f}",
            f"{commande.montant_tva:.2f}",
            f"{commande.total:.2f}",
            commande.cree_par.username if commande.cree_par else ''
        ])
    
    return response


def exporter_commandes_achat_pdf(commandes):
    """Exporter les commandes d'achat en format PDF"""
    from datetime import datetime
    
    buffer = BytesIO()
    
    # Créer le document PDF en paysage
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title_style = styles['Heading1']
    title_style.alignment = 1  # Centré
    title = Paragraph(f"Liste des Commandes d'Achat - {datetime.now().strftime('%d/%m/%Y')}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # Préparer les données du tableau
    data = [['N° Commande', 'Fournisseur', 'Date Commande', 'Date Livraison', 'Statut', 'Total']]
    
    for commande in commandes:
        data.append([
            commande.numero_commande,
            commande.fournisseur.nom[:25],  # Tronquer si trop long
            commande.date_commande.strftime('%d/%m/%Y'),
            commande.date_livraison_prevue.strftime('%d/%m/%Y') if commande.date_livraison_prevue else '',
            commande.get_statut_display(),
            f"{commande.total:,.0f} FCFA"
        ])
    
    # Ajouter ligne de total
    total_general = sum(c.total for c in commandes)
    data.append(['', '', '', '', 'TOTAL:', f"{total_general:,.0f} FCFA"])
    
    # Créer le tableau
    table = Table(data, colWidths=[4*cm, 5*cm, 3*cm, 3*cm, 3*cm, 3.5*cm])
    
    # Style du tableau
    table.setStyle(TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Corps du tableau
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        
        # Ligne de total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E5E7EB')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('ALIGN', (5, -1), (5, -1), 'RIGHT'),
    ]))
    
    elements.append(table)
    
    # Construire le PDF
    doc.build(elements)
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f"commandes_achat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# ========== FONCTIONS D'EXPORTATION ET IMPRESSION HISTORIQUE ACHATS ==========
# À AJOUTER à la fin de achats/views.py

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.html import strip_tags
import csv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

@login_required
def exporter_historique_achats(request):
    """
    Vue pour exporter l'historique des achats en Excel, CSV ou PDF
    """
    format_export = request.GET.get('format', 'excel')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    fournisseur_id = request.GET.get('fournisseur')
    
    # Date par défaut (30 derniers jours)
    if not date_debut:
        date_debut_obj = timezone.now() - timedelta(days=30)
        date_debut = date_debut_obj.strftime('%Y-%m-%d')
    else:
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
    
    if not date_fin:
        date_fin_obj = timezone.now()
        date_fin = date_fin_obj.strftime('%Y-%m-%d')
    else:
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d')
    
    # Filtrer les commandes
    commandes = CommandeAchat.objects.all()
    
    if date_debut:
        commandes = commandes.filter(date_commande__gte=date_debut)
    if date_fin:
        commandes = commandes.filter(date_commande__lte=date_fin)
    if fournisseur_id:
        commandes = commandes.filter(fournisseur_id=fournisseur_id)
    
    # Calculer les statistiques
    stats_globales = {
        'nombre_commandes': commandes.count(),
        'montant_total': commandes.aggregate(total=Sum('total'))['total'] or 0,
        'nombre_fournisseurs': commandes.values('fournisseur').distinct().count(),
        'montant_moyen': commandes.aggregate(moyenne=Avg('total'))['moyenne'] or 0,
    }
    
    # Répartition par statut
    stats_par_statut = []
    statuts = ['BROUILLON', 'CONFIRMEE', 'ENVOYEE', 'RECUE', 'FACTUREE', 'ANNULEE']
    
    for statut in statuts:
        commandes_statut = commandes.filter(statut=statut)
        nombre = commandes_statut.count()
        montant = commandes_statut.aggregate(total=Sum('total'))['total'] or 0
        
        if nombre > 0:
            stats_par_statut.append({
                'statut': statut,
                'nombre': nombre,
                'montant': montant
            })
    
    # Top fournisseurs
    top_fournisseurs = commandes.values(
        'fournisseur__nom'
    ).annotate(
        nombre=Count('id'),
        montant=Sum('total')
    ).order_by('-montant')[:10]
    
    # Top produits
    top_produits = LigneCommandeAchat.objects.filter(
        commande__in=commandes
    ).values(
        'produit__nom',
        'produit__code'
    ).annotate(
        quantite_totale=Sum('quantite'),
        montant_total=Sum(F('quantite') * F('prix_unitaire'))
    ).order_by('-montant_total')[:10]
    
    # Appeler la fonction appropriée selon le format
    if format_export == 'excel':
        return exporter_historique_excel(
            stats_globales, stats_par_statut, top_fournisseurs, 
            top_produits, date_debut_obj, date_fin_obj
        )
    elif format_export == 'csv':
        return exporter_historique_csv(
            stats_globales, stats_par_statut, top_fournisseurs, 
            top_produits, date_debut_obj, date_fin_obj
        )
    elif format_export == 'pdf':
        return exporter_historique_pdf(
            stats_globales, stats_par_statut, top_fournisseurs, 
            top_produits, date_debut_obj, date_fin_obj
        )
    else:
        return HttpResponse("Format non supporté", status=400)


def exporter_historique_excel(stats_globales, stats_par_statut, top_fournisseurs, top_produits, date_debut, date_fin):
    """Exporter l'historique des achats en format Excel avec graphiques"""
    from datetime import datetime
    
    # Créer un workbook
    wb = Workbook()
    
    # ========== FEUILLE 1 : STATISTIQUES GLOBALES ==========
    ws_stats = wb.active
    ws_stats.title = "Statistiques Globales"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    subheader_font = Font(bold=True, size=12, color="1e293b")
    value_font = Font(size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre principal
    ws_stats['A1'] = "RAPPORT HISTORIQUE DES ACHATS"
    ws_stats['A1'].font = Font(bold=True, size=16, color="6366F1")
    ws_stats.merge_cells('A1:D1')
    ws_stats['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Période
    ws_stats['A2'] = f"Période : {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    ws_stats['A2'].font = Font(size=11, color="64748b")
    ws_stats.merge_cells('A2:D2')
    ws_stats['A2'].alignment = Alignment(horizontal="center")
    
    # Statistiques globales
    current_row = 4
    ws_stats[f'A{current_row}'] = "STATISTIQUES GLOBALES"
    ws_stats[f'A{current_row}'].font = subheader_font
    ws_stats[f'A{current_row}'].fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    ws_stats.merge_cells(f'A{current_row}:B{current_row}')
    
    current_row += 2
    stats_data = [
        ('Nombre de commandes', stats_globales['nombre_commandes']),
        ('Montant total (FCFA)', stats_globales['montant_total']),
        ('Nombre de fournisseurs', stats_globales['nombre_fournisseurs']),
        ('Montant moyen (FCFA)', stats_globales['montant_moyen']),
    ]
    
    for label, value in stats_data:
        ws_stats[f'A{current_row}'] = label
        ws_stats[f'A{current_row}'].font = value_font
        ws_stats[f'A{current_row}'].border = border
        
        ws_stats[f'B{current_row}'] = value
        ws_stats[f'B{current_row}'].font = Font(bold=True, size=11)
        ws_stats[f'B{current_row}'].border = border
        ws_stats[f'B{current_row}'].alignment = Alignment(horizontal="right")
        
        if 'Montant' in label:
            ws_stats[f'B{current_row}'].number_format = '#,##0.00'
        
        current_row += 1
    
    # Ajuster les largeurs
    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 20
    
    # ========== FEUILLE 2 : RÉPARTITION PAR STATUT ==========
    ws_statut = wb.create_sheet("Répartition par Statut")
    
    # En-têtes
    headers = ['Statut', 'Nombre de commandes', 'Montant total (FCFA)', 'Pourcentage (%)']
    for col, header in enumerate(headers, start=1):
        cell = ws_statut.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Données
    total_commandes = stats_globales['nombre_commandes']
    for row_idx, stat in enumerate(stats_par_statut, start=2):
        pourcentage = (stat['nombre'] / total_commandes * 100) if total_commandes > 0 else 0
        
        data_row = [
            stat['statut'],
            stat['nombre'],
            float(stat['montant']),
            pourcentage
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws_statut.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            if col_idx == 3:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 4:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right")
    
    # Largeurs des colonnes
    ws_statut.column_dimensions['A'].width = 20
    ws_statut.column_dimensions['B'].width = 20
    ws_statut.column_dimensions['C'].width = 25
    ws_statut.column_dimensions['D'].width = 18
    
    # Ajouter un graphique en barres
    chart = BarChart()
    chart.title = "Répartition des commandes par statut"
    chart.x_axis.title = "Statut"
    chart.y_axis.title = "Nombre de commandes"
    
    data = Reference(ws_statut, min_col=2, min_row=1, max_row=len(stats_par_statut)+1)
    cats = Reference(ws_statut, min_col=1, min_row=2, max_row=len(stats_par_statut)+1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    
    ws_statut.add_chart(chart, f"F2")
    
    # ========== FEUILLE 3 : TOP FOURNISSEURS ==========
    ws_fournisseurs = wb.create_sheet("Top Fournisseurs")
    
    headers = ['Rang', 'Fournisseur', 'Nombre de commandes', 'Montant total (FCFA)']
    for col, header in enumerate(headers, start=1):
        cell = ws_fournisseurs.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    for row_idx, fournisseur in enumerate(top_fournisseurs, start=2):
        # Médailles pour le top 3
        if row_idx == 2:
            rang = "🥇 1"
        elif row_idx == 3:
            rang = "🥈 2"
        elif row_idx == 4:
            rang = "🥉 3"
        else:
            rang = str(row_idx - 1)
        
        data_row = [
            rang,
            fournisseur['fournisseur__nom'] or 'Fournisseur',
            fournisseur['nombre'],
            float(fournisseur['montant'])
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws_fournisseurs.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            if col_idx == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
    
    ws_fournisseurs.column_dimensions['A'].width = 10
    ws_fournisseurs.column_dimensions['B'].width = 35
    ws_fournisseurs.column_dimensions['C'].width = 22
    ws_fournisseurs.column_dimensions['D'].width = 25
    
    # ========== FEUILLE 4 : TOP PRODUITS ==========
    ws_produits = wb.create_sheet("Top Produits")
    
    headers = ['Rang', 'Code', 'Produit', 'Quantité totale', 'Montant total (FCFA)']
    for col, header in enumerate(headers, start=1):
        cell = ws_produits.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    for row_idx, produit in enumerate(top_produits, start=2):
        # Médailles pour le top 3
        if row_idx == 2:
            rang = "🥇 1"
        elif row_idx == 3:
            rang = "🥈 2"
        elif row_idx == 4:
            rang = "🥉 3"
        else:
            rang = str(row_idx - 1)
        
        data_row = [
            rang,
            produit['produit__code'] or '-',
            produit['produit__nom'] or 'Produit',
            produit['quantite_totale'],
            float(produit['montant_total'])
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws_produits.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            if col_idx == 5:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
    
    ws_produits.column_dimensions['A'].width = 10
    ws_produits.column_dimensions['B'].width = 15
    ws_produits.column_dimensions['C'].width = 35
    ws_produits.column_dimensions['D'].width = 20
    ws_produits.column_dimensions['E'].width = 25
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Préparer la réponse
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"historique_achats_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def exporter_historique_csv(stats_globales, stats_par_statut, top_fournisseurs, top_produits, date_debut, date_fin):
    """Exporter l'historique des achats en format CSV"""
    from datetime import datetime
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"historique_achats_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Ajouter BOM UTF-8 pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # En-tête
    writer.writerow([f"RAPPORT HISTORIQUE DES ACHATS"])
    writer.writerow([f"Période : {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"])
    writer.writerow([])
    
    # Statistiques globales
    writer.writerow(['STATISTIQUES GLOBALES'])
    writer.writerow(['Indicateur', 'Valeur'])
    writer.writerow(['Nombre de commandes', stats_globales['nombre_commandes']])
    writer.writerow(['Montant total (FCFA)', f"{stats_globales['montant_total']:.2f}"])
    writer.writerow(['Nombre de fournisseurs', stats_globales['nombre_fournisseurs']])
    writer.writerow(['Montant moyen (FCFA)', f"{stats_globales['montant_moyen']:.2f}"])
    writer.writerow([])
    
    # Répartition par statut
    writer.writerow(['RÉPARTITION PAR STATUT'])
    writer.writerow(['Statut', 'Nombre de commandes', 'Montant total (FCFA)'])
    for stat in stats_par_statut:
        writer.writerow([
            stat['statut'],
            stat['nombre'],
            f"{stat['montant']:.2f}"
        ])
    writer.writerow([])
    
    # Top fournisseurs
    writer.writerow(['TOP 10 FOURNISSEURS'])
    writer.writerow(['Rang', 'Fournisseur', 'Nombre de commandes', 'Montant total (FCFA)'])
    for idx, fournisseur in enumerate(top_fournisseurs, start=1):
        writer.writerow([
            idx,
            fournisseur['fournisseur__nom'],
            fournisseur['nombre'],
            f"{fournisseur['montant']:.2f}"
        ])
    writer.writerow([])
    
    # Top produits
    writer.writerow(['TOP 10 PRODUITS'])
    writer.writerow(['Rang', 'Code', 'Produit', 'Quantité totale', 'Montant total (FCFA)'])
    for idx, produit in enumerate(top_produits, start=1):
        writer.writerow([
            idx,
            produit['produit__code'] or '-',
            produit['produit__nom'],
            produit['quantite_totale'],
            f"{produit['montant_total']:.2f}"
        ])
    
    return response


def exporter_historique_pdf(stats_globales, stats_par_statut, top_fournisseurs, top_produits, date_debut, date_fin):
    """Exporter l'historique des achats en format PDF professionnel"""
    from datetime import datetime
    
    buffer = BytesIO()
    
    # Créer le document PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2.5*cm,
        bottomMargin=2*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Styles personnalisés
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#6366f1'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=12,
        spaceBefore=20,
        fontName='Helvetica-Bold'
    )
    
    # Titre
    elements.append(Paragraph("RAPPORT HISTORIQUE DES ACHATS", title_style))
    elements.append(Paragraph(
        f"Période : {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}", 
        subtitle_style
    ))
    
    # ========== STATISTIQUES GLOBALES ==========
    elements.append(Paragraph("Statistiques Globales", section_style))
    
    stats_data = [
        ['Indicateur', 'Valeur'],
        ['Nombre de commandes', f"{stats_globales['nombre_commandes']}"],
        ['Montant total', f"{stats_globales['montant_total']:,.0f} FCFA"],
        ['Nombre de fournisseurs', f"{stats_globales['nombre_fournisseurs']}"],
        ['Montant moyen', f"{stats_globales['montant_moyen']:,.0f} FCFA"],
    ]
    
    stats_table = Table(stats_data, colWidths=[10*cm, 6*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica'),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
    ]))
    
    elements.append(stats_table)
    elements.append(Spacer(1, 0.8*cm))
    
    # ========== RÉPARTITION PAR STATUT ==========
    if stats_par_statut:
        elements.append(Paragraph("Répartition par Statut", section_style))
        
        statut_data = [['Statut', 'Nombre', 'Montant total (FCFA)']]
        for stat in stats_par_statut:
            statut_data.append([
                stat['statut'],
                str(stat['nombre']),
                f"{stat['montant']:,.0f} FCFA"
            ])
        
        statut_table = Table(statut_data, colWidths=[6*cm, 4*cm, 6*cm])
        statut_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(statut_table)
        elements.append(Spacer(1, 0.8*cm))
    
    # ========== TOP FOURNISSEURS ==========
    if top_fournisseurs:
        elements.append(Paragraph("Top 10 Fournisseurs", section_style))
        
        fournisseurs_data = [['Rang', 'Fournisseur', 'Commandes', 'Montant (FCFA)']]
        for idx, fournisseur in enumerate(top_fournisseurs, start=1):
            # Médailles
            if idx == 1:
                rang = "🥇 1"
            elif idx == 2:
                rang = "🥈 2"
            elif idx == 3:
                rang = "🥉 3"
            else:
                rang = str(idx)
            
            fournisseurs_data.append([
                rang,
                fournisseur['fournisseur__nom'][:35],
                str(fournisseur['nombre']),
                f"{fournisseur['montant']:,.0f}"
            ])
        
        fournisseurs_table = Table(fournisseurs_data, colWidths=[2*cm, 7*cm, 3*cm, 4*cm])
        fournisseurs_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(fournisseurs_table)
        elements.append(PageBreak())
    
    # ========== TOP PRODUITS ==========
    if top_produits:
        elements.append(Paragraph("Top 10 Produits Achetés", section_style))
        
        produits_data = [['Rang', 'Code', 'Produit', 'Qté', 'Montant (FCFA)']]
        for idx, produit in enumerate(top_produits, start=1):
            # Médailles
            if idx == 1:
                rang = "🥇 1"
            elif idx == 2:
                rang = "🥈 2"
            elif idx == 3:
                rang = "🥉 3"
            else:
                rang = str(idx)
            
            produits_data.append([
                rang,
                produit['produit__code'] or '-',
                produit['produit__nom'][:30],
                str(produit['quantite_totale']),
                f"{produit['montant_total']:,.0f}"
            ])
        
        produits_table = Table(produits_data, colWidths=[2*cm, 2.5*cm, 6*cm, 2.5*cm, 3*cm])
        produits_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0ea5e9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(produits_table)
    
    # Pied de page
    elements.append(Spacer(1, 1.5*cm))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#94a3b8'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph(
        f"Rapport généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        footer_style
    ))
    
    # Construire le PDF
    doc.build(elements)
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f"historique_achats_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# ========== FONCTIONS D'IMPRESSION ET TÉLÉCHARGEMENT HISTORIQUE ACHATS - VERSION CORRIGÉE ==========
# À AJOUTER à la fin de achats/views.py (après toutes les autres fonctions)

from reportlab.platypus import PageBreak

@login_required
def imprimer_historique_achats(request):
    """
    Vue pour afficher une version imprimable de l'historique des achats
    """
    # Récupérer les paramètres de filtrage
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    fournisseur_id = request.GET.get('fournisseur')
    
    # Date par défaut (30 derniers jours)
    if not date_debut:
        date_debut = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_fin:
        date_fin = timezone.now().strftime('%Y-%m-%d')
    
    # Filtrer les commandes
    commandes = CommandeAchat.objects.all()
    
    if date_debut:
        commandes = commandes.filter(date_commande__gte=date_debut)
    if date_fin:
        commandes = commandes.filter(date_commande__lte=date_fin)
    if fournisseur_id:
        commandes = commandes.filter(fournisseur_id=fournisseur_id)
    
    # Statistiques globales
    stats_globales = {
        'nombre_commandes': commandes.count(),
        'montant_total': commandes.aggregate(total=Sum('total'))['total'] or Decimal('0'),
        'nombre_fournisseurs': commandes.values('fournisseur').distinct().count(),
        'montant_moyen': commandes.aggregate(moyenne=Avg('total'))['moyenne'] or Decimal('0'),
    }
    
    # Répartition par statut
    stats_par_statut = []
    statuts = ['BROUILLON', 'CONFIRMEE', 'ENVOYEE', 'RECUE', 'FACTUREE', 'ANNULEE']
    
    statut_labels = {
        'BROUILLON': 'Brouillon',
        'CONFIRMEE': 'Confirmée',
        'ENVOYEE': 'Envoyée',
        'RECUE': 'Reçue',
        'FACTUREE': 'Facturée',
        'ANNULEE': 'Annulée'
    }
    
    for statut in statuts:
        commandes_statut = commandes.filter(statut=statut)
        nombre = commandes_statut.count()
        montant = commandes_statut.aggregate(total=Sum('total'))['total'] or Decimal('0')
        pourcentage = (nombre / stats_globales['nombre_commandes'] * 100) if stats_globales['nombre_commandes'] > 0 else 0
        
        if nombre > 0:
            stats_par_statut.append({
                'statut': statut,
                'statut_display': statut_labels.get(statut, statut),
                'nombre': nombre,
                'montant': float(montant),
                'pourcentage': float(pourcentage)
            })
    
    # Top 10 fournisseurs
    top_fournisseurs = list(commandes.values(
        'fournisseur__nom'
    ).annotate(
        nombre=Count('id'),
        montant=Sum('total')
    ).order_by('-montant')[:10])
    
    # Convertir les Decimal en float pour éviter les problèmes de sérialisation
    for f in top_fournisseurs:
        if f.get('montant'):
            f['montant'] = float(f['montant'])
    
    # Top 10 produits
    top_produits = list(LigneCommandeAchat.objects.filter(
        commande__in=commandes
    ).values(
        'produit__nom',
        'produit__code'
    ).annotate(
        quantite_totale=Sum('quantite'),
        montant_total=Sum(F('quantite') * F('prix_unitaire'))
    ).order_by('-montant_total')[:10])
    
    # Convertir les Decimal en float
    for p in top_produits:
        if p.get('montant_total'):
            p['montant_total'] = float(p['montant_total'])
    
    context = {
        'stats_globales': {
            'nombre_commandes': stats_globales['nombre_commandes'],
            'montant_total': float(stats_globales['montant_total']),
            'nombre_fournisseurs': stats_globales['nombre_fournisseurs'],
            'montant_moyen': float(stats_globales['montant_moyen']),
        },
        'stats_par_statut': stats_par_statut,
        'top_fournisseurs': top_fournisseurs,
        'top_produits': top_produits,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'date_impression': timezone.now().strftime('%d/%m/%Y à %H:%M'),
    }
    
    return render(request, 'achats/imprimer_historique.jinja', context)


@login_required
def telecharger_historique_pdf(request):
    """
    Vue pour télécharger l'historique des achats en PDF professionnel
    """
    from datetime import datetime
    
    # Récupérer les paramètres de filtrage
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    fournisseur_id = request.GET.get('fournisseur')
    
    # Date par défaut (30 derniers jours)
    if not date_debut:
        date_debut_obj = timezone.now() - timedelta(days=30)
        date_debut = date_debut_obj.strftime('%Y-%m-%d')
    else:
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
    
    if not date_fin:
        date_fin_obj = timezone.now()
        date_fin = date_fin_obj.strftime('%Y-%m-%d')
    else:
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d')
    
    # Filtrer les commandes
    commandes = CommandeAchat.objects.all()
    
    if date_debut:
        commandes = commandes.filter(date_commande__gte=date_debut)
    if date_fin:
        commandes = commandes.filter(date_commande__lte=date_fin)
    if fournisseur_id:
        commandes = commandes.filter(fournisseur_id=fournisseur_id)
    
    # Statistiques globales
    stats_globales = {
        'nombre_commandes': commandes.count(),
        'montant_total': commandes.aggregate(total=Sum('total'))['total'] or Decimal('0'),
        'nombre_fournisseurs': commandes.values('fournisseur').distinct().count(),
        'montant_moyen': commandes.aggregate(moyenne=Avg('total'))['moyenne'] or Decimal('0'),
    }
    
    # Répartition par statut
    stats_par_statut = []
    statuts = ['BROUILLON', 'CONFIRMEE', 'ENVOYEE', 'RECUE', 'FACTUREE', 'ANNULEE']
    
    statut_labels = {
        'BROUILLON': 'Brouillon',
        'CONFIRMEE': 'Confirmée',
        'ENVOYEE': 'Envoyée',
        'RECUE': 'Reçue',
        'FACTUREE': 'Facturée',
        'ANNULEE': 'Annulée'
    }
    
    for statut in statuts:
        commandes_statut = commandes.filter(statut=statut)
        nombre = commandes_statut.count()
        montant = commandes_statut.aggregate(total=Sum('total'))['total'] or Decimal('0')
        
        if nombre > 0:
            stats_par_statut.append({
                'statut': statut_labels.get(statut, statut),
                'nombre': nombre,
                'montant': montant
            })
    
    # Top 10 fournisseurs
    top_fournisseurs = commandes.values(
        'fournisseur__nom'
    ).annotate(
        nombre=Count('id'),
        montant=Sum('total')
    ).order_by('-montant')[:10]
    
    # Top 10 produits
    top_produits = LigneCommandeAchat.objects.filter(
        commande__in=commandes
    ).values(
        'produit__nom',
        'produit__code'
    ).annotate(
        quantite_totale=Sum('quantite'),
        montant_total=Sum(F('quantite') * F('prix_unitaire'))
    ).order_by('-montant_total')[:10]
    
    # Créer le PDF avec ReportLab
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Style personnalisé pour le titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#6366F1'),
        spaceAfter=20,
        alignment=1,  # Centré
        fontName='Helvetica-Bold'
    )
    
    # Style pour les sous-titres
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=12,
        spaceBefore=20,
        fontName='Helvetica-Bold'
    )
    
    # Style pour le texte normal
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#475569'),
    )
    
    # === EN-TÊTE ===
    title = Paragraph("RAPPORT D'HISTORIQUE DES ACHATS", title_style)
    elements.append(title)
    
    # Période
    periode_text = f"Période du {date_debut_obj.strftime('%d/%m/%Y')} au {date_fin_obj.strftime('%d/%m/%Y')}"
    periode = Paragraph(periode_text, normal_style)
    elements.append(periode)
    elements.append(Spacer(1, 0.3*cm))
    
    # Date de génération
    date_generation = Paragraph(
        f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}",
        normal_style
    )
    elements.append(date_generation)
    elements.append(Spacer(1, 1*cm))
    
    # === STATISTIQUES GLOBALES ===
    elements.append(Paragraph("Statistiques Globales", subtitle_style))
    
    stats_data = [
        ['Indicateur', 'Valeur'],
        ['Nombre de commandes', f"{stats_globales['nombre_commandes']:,}"],
        ['Montant total', f"{float(stats_globales['montant_total']):,.0f} FCFA"],
        ['Nombre de fournisseurs', f"{stats_globales['nombre_fournisseurs']:,}"],
        ['Montant moyen/commande', f"{float(stats_globales['montant_moyen']):,.0f} FCFA"],
    ]
    
    stats_table = Table(stats_data, colWidths=[10*cm, 6*cm])
    stats_table.setStyle(TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Corps
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    
    elements.append(stats_table)
    elements.append(Spacer(1, 1*cm))
    
    # === RÉPARTITION PAR STATUT ===
    if stats_par_statut:
        elements.append(Paragraph("Répartition par Statut", subtitle_style))
        
        statut_data = [['Statut', 'Nombre', 'Montant Total']]
        
        for stat in stats_par_statut:
            statut_data.append([
                stat['statut'],
                f"{stat['nombre']:,}",
                f"{float(stat['montant']):,.0f} FCFA"
            ])
        
        # Ligne de total
        total_commandes = sum(s['nombre'] for s in stats_par_statut)
        total_montant = sum(float(s['montant']) for s in stats_par_statut)
        statut_data.append(['TOTAL', f"{total_commandes:,}", f"{total_montant:,.0f} FCFA"])
        
        statut_table = Table(statut_data, colWidths=[8*cm, 4*cm, 4*cm])
        statut_table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Corps
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8FAFC')]),
            
            # Ligne de total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E5E7EB')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(statut_table)
        elements.append(Spacer(1, 0.5*cm))
    else:
        elements.append(Paragraph("Répartition par Statut", subtitle_style))
        elements.append(Paragraph("Aucune commande trouvée pour cette période.", normal_style))
        elements.append(Spacer(1, 0.5*cm))
    
    # === NOUVELLE PAGE ===
    elements.append(PageBreak())
    
    # === TOP FOURNISSEURS ===
    elements.append(Paragraph("Top 10 Fournisseurs", subtitle_style))
    
    if top_fournisseurs and len(list(top_fournisseurs)) > 0:
        fournisseurs_data = [['Rang', 'Fournisseur', 'Commandes', 'Montant Total']]
        
        for idx, f in enumerate(top_fournisseurs, start=1):
            fournisseurs_data.append([
                str(idx),
                f['fournisseur__nom'][:30] if f.get('fournisseur__nom') else 'N/A',
                f"{f['nombre']:,}",
                f"{float(f['montant']):,.0f} FCFA"
            ])
        
        # Vérifier qu'il y a au moins une ligne de données
        if len(fournisseurs_data) > 1:
            fournisseurs_table = Table(fournisseurs_data, colWidths=[2*cm, 7*cm, 3*cm, 4*cm])
            fournisseurs_table.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Corps
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ]))
            
            # Mise en évidence du top 3 seulement s'il y a assez de lignes
            if len(fournisseurs_data) > 1:
                fournisseurs_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#FEF3C7')),
                ]))
            if len(fournisseurs_data) > 2:
                fournisseurs_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#E5E7EB')),
                ]))
            if len(fournisseurs_data) > 3:
                fournisseurs_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#FED7AA')),
                ]))
            
            elements.append(fournisseurs_table)
        else:
            elements.append(Paragraph("Aucun fournisseur trouvé pour cette période.", normal_style))
    else:
        elements.append(Paragraph("Aucun fournisseur trouvé pour cette période.", normal_style))
    
    elements.append(Spacer(1, 1*cm))
    
    # === TOP PRODUITS ===
    elements.append(Paragraph("Top 10 Produits Achetés", subtitle_style))
    
    if top_produits and len(list(top_produits)) > 0:
        produits_data = [['Rang', 'Produit', 'Code', 'Quantité', 'Montant']]
        
        for idx, p in enumerate(top_produits, start=1):
            produits_data.append([
                str(idx),
                p['produit__nom'][:25] if p.get('produit__nom') else 'N/A',
                p['produit__code'] if p.get('produit__code') else '-',
                f"{p['quantite_totale']:,}" if p.get('quantite_totale') else '0',
                f"{float(p['montant_total']):,.0f} FCFA" if p.get('montant_total') else '0 FCFA'
            ])
        
        # Vérifier qu'il y a au moins une ligne de données
        if len(produits_data) > 1:
            produits_table = Table(produits_data, colWidths=[1.5*cm, 6*cm, 2.5*cm, 2.5*cm, 3.5*cm])
            produits_table.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0EA5E9')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Corps
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ]))
            
            elements.append(produits_table)
        else:
            elements.append(Paragraph("Aucun produit trouvé pour cette période.", normal_style))
    else:
        elements.append(Paragraph("Aucun produit trouvé pour cette période.", normal_style))
    
    # === PIED DE PAGE ===
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#94A3B8'),
        alignment=1  # Centré
    )
    
    footer_text = f"<i>Rapport généré automatiquement le {timezone.now().strftime('%d/%m/%Y à %H:%M')} - Système ERP</i>"
    footer = Paragraph(footer_text, footer_style)
    elements.append(Spacer(1, 2*cm))
    elements.append(footer)
    
    # Construire le PDF
    doc.build(elements)
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f"historique_achats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response